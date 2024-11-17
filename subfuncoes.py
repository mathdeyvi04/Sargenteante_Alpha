# Devemos ter a função da limpeza
import os
import datetime as dt
from datetime import datetime
from tkinter import messagebox
from tkinter import ttk
from tkinter import *
import openpyxl as xl
from openpyxl.styles import Font, PatternFill

limite_de_busca = 50


# Devemos ter a função que vai trocar o tipo de mes para a letra do alfabeto
def codigo_mes(mes_info, opcao=True):
    # Quer dizer que queremos obter o codigo referente ao mes
    mes_info = str(mes_info)

    if mes_info.isalpha() and len(mes_info) == 3:

        trocas = {'Jan': 'a',
                  'Fev': 'b',
                  'Mar': 'c',
                  'Apr': 'd',
                  'May': 'e',
                  'June': 'f',
                  'July': 'g',
                  'Aug': 'h',
                  'Sept': 'i',
                  'Oct': 'j',
                  'Nov': 'k',
                  'Dec': 'l'
                  }

        return trocas[mes_info]

    else:

        meses = ['a', 'b', 'c', 'd', 'e', 'f',
                 'g', 'h', 'i', 'j', 'k', 'l']

        if opcao:
            mes_info = int(mes_info)
            return meses[mes_info - 1]
        else:
            # quer dizer que vamos receber uma letra codificada e queremos o mes em numero
            for c in range(0, 12):
                if meses[c] == mes_info:
                    return c + 1


def obtendo_hoje():
    dia = dt.datetime.now().day
    mes = dt.datetime.now().month
    return f'{dia}{codigo_mes(mes)}'


def util_ou_nao(dia):
    ANO = int(dt.datetime.now().year)
    data = dt.date(ANO, int(codigo_mes(dia[-1], False)), int(dia.replace(dia[-1], '')))
    dia_c = data.weekday()

    if dia_c in [5, 6]:
        # quer dizer que é final de semana
        return False
    else:
        # é um dia da semana, devemos verificar se é feriado
        data = str(data)
        import holidays as hl

        br_holi = hl.Brazil()

        if data in br_holi:
            return False
        else:
            # se nao for final de semana nem feriado
            return True


def seguinte(codigo_longe):
    ANO = dt.datetime.now().year

    mes = codigo_mes(codigo_longe[-1], opcao=False)
    dia = int(codigo_longe.replace(codigo_longe[-1], ''))

    data_mais_longe = dt.date(ANO, mes, dia)

    # Devemos acrescentar em uma unidade para o próximo
    try:
        # Naturalmente, vai acontecer a soma, 15 + 1 = 16
        data_mais_longe = data_mais_longe.replace(day=data_mais_longe.day + 1)
    except ValueError:
        # Há casos de erros, 30 + 1 = 31 ou 1, dependendo do mes, compreende?
        # O except cuidará desses casos
        data_mais_longe = data_mais_longe.replace(month=data_mais_longe.month + 1, day=1)

    # Agora temos a data correta.
    # Basta deixá-la em codigo
    return f'{data_mais_longe.day}{codigo_mes(int(data_mais_longe.month), True)}'


def comparando_codigos(data1, data2):
    # Essa função vai decidir qual data é mais antiga,
    # Vai sempre me dizer se a data1 é mais proxima que a data2

    if data1[-1] == data2[-1]:
        # print('Estão no mesmo mês')
        # ambas datas estão no mesmo mes

        if int(data1.replace(data1[-1], '')) < int(data2.replace(data2[-1], '')):
            # print(f'Estou comparando os dias {data1.replace(data1[-1], "")} se é menor que {data2.replace(data2[-1], "")}')
            # Logo, data1 é mais antiga que data2
            return True
        else:
            return False
    else:
        # ambas estão em mses difentes, logo, podemos

        if data1[-1] < data2[-1]:
            # print('a está em um mês mais novo, logo a < b')
            return True
        else:
            # print('a não está em um mês mais novo, logo a > b')
            return False


def ordenador_escalas(escalas_desordenadas):
    escalas_desordenadas = [esc.replace('escala', '') for esc in escalas_desordenadas]
    tam = len(escalas_desordenadas)

    ordenadas = []

    # Com essa função, vamos obter a menor data na lista
    def menor():
        mais_proxima = escalas_desordenadas[0]
        for data in escalas_desordenadas:
            if comparando_codigos(mais_proxima, data):
                # Se for verdade, quer dizer que mais_proxima é a menor
                pass
            else:
                # Ou é igual ou data é menor
                if data != mais_proxima:
                    mais_proxima = data

        return mais_proxima

    while True:

        # O quebrador
        if len(ordenadas) == tam:
            break

        menor_data = menor()

        ordenadas.append(menor_data)

        # Vamos retirar para não ficarmos com um loop infinito
        escalas_desordenadas.remove(menor_data)

    return ['escala' + dat for dat in ordenadas]


def obtendo_dif(arquivo, data_hoje=''):
    if data_hoje == '':
        data_hoje = dt.date.today()

    ANO = int(dt.datetime.now().year)

    codigo_referente = arquivo.replace('escala', '')

    # estava dando erro, pois o replace so vai retornar uma string, não um inteiro
    data_referente = dt.date(ANO,
                             codigo_mes(codigo_referente[-1], opcao=False),
                             int(codigo_referente.replace(codigo_referente[-1], '')))

    # Caso o usuario solicite um caso
    if type(data_hoje) == str:
        mes = codigo_mes(data_hoje[-1], opcao=False)
        dia = int(str(data_hoje).replace(data_hoje[-1], ''))
        data_hoje = dt.date(ANO, mes, dia)

    # afinal queremos pegar o tempo até que ela ocorra
    return (data_referente - data_hoje).days


def obtendo_data(arquivo):
    arquivo = arquivo.replace('escala', '')

    mes = codigo_mes(arquivo[-1], False)
    nomes_meses = ['Janeiro',
                   'Fevereiro',
                   'Março',
                   'Abril',
                   'Maio',
                   'Junho',
                   'Julho',
                   'Agosto',
                   'Setembro',
                   'Outubro',
                   'Novembro',
                   'Decembro']
    mes = nomes_meses[mes - 1]

    dia = str(arquivo.replace(arquivo[-1], ''))

    return ' de '.join([dia, mes])


def destruidor(infor, arquivo, se_teste=False):
    # Vamos receber a lista de informações de cada turma e o arquivo em que ela se encontra.

    try:
        p = open(arquivo, 'r')
        linhas = p.readlines()
        p.close()

        presente = []
        # Vamos descobrir em quais linhas está informação está
        for linha in linhas:
            # Vamos pegar cada pessoa das informações
            for pessoa in infor:
                # Vamos verificar se essa pessoa está na linha
                if pessoa in linha:
                    # Caso esteja, devemos verificar se está linha já foi adicionada
                    if linha in presente:
                        pass
                    else:
                        # Caso não, devemos adicioná-la para ser apagada
                        presente.append(linha)

        # Caso realmente exista algo a ser apagado
        if len(presente) != 0:
            if se_teste:
                # Vamos construir aqui para mostrar ao usuário que deu certo o test4e dele

                print(f'No arquivo {arquivo} iriamos apagar alguem')

            else:
                # Vamos apagar as linhas
                for deve_ser_Apagada in presente:
                    linhas.remove(deve_ser_Apagada)

                # E agora reescrever
                os.remove(arquivo)

                primeiro = True
                with open(arquivo, 'x') as arq:
                    for linha_ in linhas:
                        if primeiro:
                            arq.write(f'{linha_}')
                            primeiro = False
                        else:
                            arq.write(f'\n{linha_}')

        else:
            # Nem estava nesse arquivo
            pass
    except:
        messagebox.showerror(message=f'Erro ao tentar apagar em {arq}')
        return None


def enviando_email(destino, texto):
    try:
        import smtplib
        from smtplib.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        def obtendo_credenciais_do_usuario():
            p = open('../SargenteanteAlpha/SubArquivos/HistoricoVisual')
            linhas = p.readlines()
            p.close()

            return linhas[-1].split(';')

        cred = obtendo_credenciais_do_usuario()

        # Informações da conta de e-mail
        remetente_email = cred[0]
        remetente_senha = cred[1]
        destinatario_email = destino

        # Configuração do servidor SMTP do Gmail
        smtp_server = 'smtp.gmail.com'
        smtp_port = 587

        # Criar objeto SMTP
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()  # Iniciar conexão segura

        # Faça login na conta de e-mail
        server.login(remetente_email, remetente_senha)

        # Criar mensagem de e-mail
        mensagem = MIMEMultipart()
        mensagem['From'] = remetente_email
        mensagem['To'] = destinatario_email
        mensagem['Subject'] = 'Escala de Serviço'

        # Corpo do e-mail
        corpo = texto
        mensagem.attach(MIMEText(corpo, 'plain'))

        # Enviar e-mail
        texto_email = mensagem.as_string()
        server.sendmail(remetente_email, destinatario_email, texto_email)

        # Encerrar a conexão SMTP
        server.quit()

    except:
        messagebox.showerror(message='Erro ao tentar enviar email')


def enviando_sms(texto1, numero1):
    from twilio.rest import Client  # Isso é para conseguirmos enviar SMS
    try:
        # Colocando credenciais, estava dando erro porque haviamos colocado na ordem trocada
        sid_ = '#'
        token_ = '#'

        # Enviando
        cliente_safo = Client(sid_, token_)
        cliente_safo.messages.create(from_='#', to='+55' + numero1,
                                     body=texto1)

        messagebox.showinfo(message='Envio Bem Sucedido')
    except Exception as e:
        messagebox.showerror(title='ERROR',
                             message='Houve um erro no envio do SMS, muito'
                                     'provavel que seja devido ao número não veri-'
                                     'ficado, vá ao Twilio e cadastre o número')

        print(e)


def historicando(contato, opcao):
    if opcao:
        try:
            # queremos salvar
            # Primeiro, devemos obter os dados existentes
            p = open('../SargenteanteAlpha/SubArquivos/HistoricoVisual')
            linhas = list(p.readlines())
            # Sempre vai ser o primeiro que vai ter um '\n'
            p.close()

            index = 0 if contato.isnumeric() else 1

            linhas[index] = linhas[index].replace("\n", '')
            linhas[index] = linhas[index].split(';')
            linhas[index].append(f'{contato}\n')
            linhas[index] = ';'.join(linhas[index])

            # Agora salvando
            os.remove('../SargenteanteAlpha/SubArquivos/HistoricoVisual')

            with open("../SargenteanteAlpha/SubArquivos/HistoricoVisual", 'x') as hv:
                for linha in linhas:
                    hv.write(linha)
        except:
            messagebox.showerror(message='Erro ao tentar salvar')

    else:
        try:
            # queremos ler os dados

            num = int(contato)

            i = 0
            with open('../SargenteanteAlpha/SubArquivos/HistoricoVisual') as hv:
                for linha in hv:
                    # Vamos varrer e procurar os dados que queremos
                    i += 1

                    if i == num:
                        # chegamos ao que queriamos
                        if '\n' in linha:
                            linha = linha.replace('\n', '')
                        linha = linha.split(";")
                        return linha

        except:
            messagebox.showerror(message='Erro ao tentar pegar lista de contatos')


def limpeza(escala, decisao):
    """"""
    # Quando formos apagar uma escala, devemos escrever ela na planilha do Sanha
    # Devemos fazer ela vasculhar por si só também a pasta das escalas e verificar se há algumas

    '''Como vamos resolver o sanha dos reservas no registro?'''

    try:
        def escrevendo(apagada):
            # Vai receber uma escala por vez

            guarnicao = []
            para_planilha = []

            def pegando():
                # Apenas precisamos guardar a informação do nome da pessoa

                with open(rf'Escalas/{apagada}') as esc:
                    for linha_ in esc:
                        linha = linha_.split("-")
                        completo = pegando_info(linha[0], [0, 1])
                        guarnicao.append(completo)
                        para_planilha.append([completo, serv(linha[1], False)])

            pegando()

            # Daqui, ja conseguimos obter apenas o nome

            # devemos pegar a data atual sem codigo
            data_hoje = dt.date.today()

            try:

                def escreve_planilha():
                    # vamos ter um .txt indicando o nome das planilhas e as suas últimas colunas usadas
                    ANO = dt.datetime.now().year

                    try:
                        # agora devemos abrir o arquivo .txt citado e verificar a ultima coluna usada em lixeira{ANO}^.xlsx
                        def obtendo_coluna_a_ser_preenchida():

                            def ultima_coluna_preenchida():
                                try:
                                    with open(f'Lixeira/identacao', 'r') as id1:
                                        for linha in id1:
                                            if f'lixeira{ANO}' in linha:
                                                if '\n' in linha:
                                                    linha = linha.replace('\n', '')
                                                linha = linha.split('=')
                                                return linha[1]

                                    # Caso passemos, quer dizer que o arquivo vai ser criado
                                    # Sendo assim
                                    return None
                                except FileNotFoundError:
                                    # Vamos criar a sanha aqui mesmo.
                                    p = open('Lixeira/identacao', 'x')
                                    p.close()
                                    return None

                            ultima_col = ultima_coluna_preenchida()

                            # Caso seja a primeira vez
                            if ultima_col is None:
                                # Deve ser esse return mesmo
                                return 'A'

                            # Devemos obter a proxima coluna
                            # A lógica é A, B, ..., Z, AA, AB, ..., AZ, BA, ... ..., ZZ, AAA, AAB,...
                            def proxima(ultima):

                                ultima = list(ultima)

                                # Há um Sanha que pode ser dividido
                                if ultima[-1] == 'Z':
                                    # quer dizer que vamos para uma classe maior

                                    # Necessariamente
                                    ultima[-1] = 'A'

                                    # E o aumento de classe
                                    ultima[-2] = chr(ord(ultima[-2]) + 1)
                                else:

                                    # o Sanha continua igual, logo não teremos mudança de base
                                    ultima[-1] = chr(ord(ultima[-1]) + 1)

                                return ''.join(ultima)

                            proxima_col = proxima(ultima_col)

                            # Sendo assim, vamos devolver
                            return proxima_col

                        coluna = obtendo_coluna_a_ser_preenchida()

                    except:
                        messagebox.showerror(message='Erro ao tentar obter coluna a ser preenchida')

                    try:
                        try:
                            # De posse da última coluna a ser preenchida, vamos mexer no Sanha

                            def preenchendo():
                                nome = fr'Lixeira/lixeira{ANO}.xlsx'

                                def situacao_planilha():

                                    if coluna == 'A':
                                        # Se for a primeira vez
                                        return xl.Workbook()
                                    else:
                                        # Caso não seja a primeira
                                        return xl.load_workbook(nome)

                                lixo = situacao_planilha()

                                # Vamos deixar o Sanha pronto para ser modificado
                                pagina = lixo.active

                                # Escrevendo o titulo
                                info1 = obtendo_data(apagada)
                                pagina[coluna + '1'] = info1
                                pagina[coluna + '1'].font = Font(bold=True)

                                i = 2
                                for completo, tipo in para_planilha:
                                    info = completo + ' >> ' + tipo
                                    pagina.column_dimensions[coluna].width = len(info) + 3
                                    pagina[coluna + f'{i}'] = info
                                    i += 1

                                lixo.save(nome)

                            preenchendo()
                        except:
                            messagebox.showerror(message='Erro ao tentar preencher a planilha')

                        try:
                            # Se conseguimos, vamos atualizar a identação
                            def atualizando_identacao():

                                aberto = open("Lixeira/identacao", 'r')
                                linhas = aberto.readlines()
                                aberto.close()

                                c = 0
                                for linha in linhas:
                                    if f'lixeira{ANO}' in linha:
                                        # Se acharmos
                                        linha = linha.split("=")
                                        if '\n' in linha[1]:
                                            linha[1] = coluna + '\n'
                                        else:
                                            linha[1] = coluna

                                        linhas[c] = '='.join(linha)

                                        # Note que nehuma das posteriores foi modificada, logo podemos continuar daqui
                                        # apagando
                                        os.remove("Lixeira/identacao")

                                        # Escrevendo o sanha
                                        with open("Lixeira/identacao", 'x') as id1:
                                            for linha1 in linhas:
                                                id1.write(linha1)

                                        # achamos o sanha
                                        return 0

                                    c += 1

                                # Caso esse ano não esteja presente
                                with open("Lixeira/identacao", 'a') as id2:
                                    if len(linhas) == 0:
                                        id2.write(f'lixeira{ANO}={coluna}')
                                    else:
                                        id2.write(f'\nlixeira{ANO}={coluna}')

                            atualizando_identacao()
                        except:
                            messagebox.showerror(message='Erro ao tentar atualizar última coluna')

                    except:
                        messagebox.showerror(message='Erro ao tentar atualizar planilha')

                escreve_planilha()

            except:
                messagebox.showerror(message='Erro ao salvar na planilha de lixeira')

            try:
                def escreve_registro():
                    # vai escrever dinamicamente nos registro quando é a última data do serviço
                    # ja temos as pessoas que devem ter a data alterada
                    # então bastará encontrá-las

                    # Obtendo a lista geral
                    houve_mudanca = []  # vai salvar os arquivos que fizemos mudanças

                    def pegando_lista():
                        todes = {}

                        for arq in os.listdir('../SargenteanteAlpha/Registros'):
                            # vai pegar todos em cada registro
                            todes[arq] = []
                            houve_mudanca.append(arq)
                            with open(f'Registros/{arq}', 'r') as reg:
                                for linha_ in reg:
                                    linha = linha_.split(",")
                                    todes[arq].append(linha)

                        return todes

                    todos = pegando_lista()

                    # Verificando as pessoas e realizando as trocas
                    def procurando_realizando():

                        for militar in guarnicao:
                            for arquiv in houve_mudanca:
                                c = 0
                                mudou = False
                                for pessoa_completo, ultimo in todos[arquiv]:
                                    if pessoa_completo == militar:
                                        # achamos a pessoa que estava de serviço na lista de registro da respectiva turma
                                        # devemos saber se é o último
                                        if todos[arquiv][-1][0] == pessoa_completo:
                                            todos[arquiv][-1][1] = str(data_hoje)
                                        else:
                                            todos[arquiv][c][1] = str(data_hoje) + '\n'

                                        mudou = True

                                    c += 1

                                if mudou is False:
                                    houve_mudanca.remove(arquiv)

                    procurando_realizando()

                    # vamos escrever
                    def digitando():
                        for arq in houve_mudanca:
                            os.remove(f'Registros/{arq}')

                            with open(f'Registros/{arq}', 'x') as reg:
                                for sublista in todos[arq]:
                                    linha = ','.join(sublista)
                                    reg.write(linha)

                    digitando()

                escreve_registro()
            except:
                messagebox.showerror(message='Erro ao fazer modificações nos registros.')

        if decisao == 0:
            # quer dizer que vamos apenas pegar a escala e apaga-la safamente.
            escrevendo(escala)
            os.remove(f"Escalas/{escala}")
        elif decisao != 0 and escala is None:
            hoje = obtendo_hoje()

            # vamos vasculhar e achar as inconsistencias
            for arquivo in os.listdir('../SargenteanteAlpha/Escalas'):
                if comparando_codigos(arquivo.replace('escala', ''), hoje):
                    # Se existir uma mais antiga, deve ser apagada
                    escrevendo(arquivo)
                    os.remove(f'Escalas/{arquivo}')
        else:
            pass

    except:
        messagebox.showerror(message='Erro ao tentar realizar limpeza')


def serv(carac_serv='', crip=False):
    # O padrão para as configurações será x-yzM
    # onde x é o codigo do serviço, yz representam os anos que realizam esse serviço e M é o sexo, obvio.

    carac_serv = carac_serv.replace('\n', '')
    tipos = {
        "1": 'PtSegMasc',
        "1.2": 'PtSegFem',
        "1.4": '3° Piso',
        "1.8": 'Cabo',
        "2": 'Sentinela',
        "3": 'Cabo da Guarda',
        "3.5": 'Cdt da Guarda',
        "4": 'Sgt',
        "5": 'Aux Of'
    }

    if crip is False:
        return tipos[carac_serv]
    else:
        # Quer dizer que queremos criptografar
        for key, chave in tipos.items():
            if chave == carac_serv:
                return key


def informador(mestre, posx, posy, mensagem, texto=' i '):
    def info():
        # mostra uma caixa de texto
        messagebox.showinfo(title='INFO', message=mensagem)

    Button(mestre, text=texto, command=info).place(x=posx, y=posy + 3, height=20)


def criando_janela(titulo, comp, alt):
    janela_ = Tk()
    janela_.title(titulo)
    janela_.geometry(f'{comp}x{alt}')
    janela_.configure(bg="#dde")
    janela_.resizable(False, False)

    return janela_


def pegando_info(info, entrada_saida, arquivo=None, count=False):
    def busca_logica(numero):

        turma_logica = str(int(numero[:2]) + 4)
        # Vamos abrir a turma lógica da pessoa
        try:
            with open(f'Arquivos/CFG-{turma_logica}.txt', 'r') as base:
                for linha in base:
                    linha = linha.split("-")
                    if linha[0] != '\n':
                        if numero == linha[0]:
                            # achamos o sanhudo
                            if entrada_saida[-1] == -1:
                                # so queremos saber se existe e em qual arquivo
                                return True, turma_logica
                            else:
                                return True, linha[entrada_saida[1]]
            # Se chegar aqui, não achou o sanhudo
            return turma_logica
        except FileNotFoundError:
            # Essa base não existe
            return turma_logica

    def busca_burra():
        for arq in os.listdir('../SargenteanteAlpha/Arquivos'):
            if '.txt' in arq:
                with open(f'Arquivos/{arq}') as base:
                    for linha in base:
                        linha = linha.split('-')
                        if linha[0] != '\n':
                            if linha[entrada_saida[0]] == info:
                                # achamos o guerreiro na base
                                if entrada_saida[1] == -1:
                                    # So queremos verificar se ele existe em alguma base de dados
                                    arq = arq.replace(".txt", '')
                                    arq = arq.replace('Arquivos/CFG-', '')

                                    return True, arq
                                else:
                                    return True, linha[entrada_saida[1]]

        return [False, 0]

    def busca_semiburra(arquivo1):
        if 'Arquivos' not in arquivo1:
            arquivo1 = 'Arquivos/' + arquivo1
        with open(arquivo1) as g:
            for linha in g:
                linha = linha.split('-')
                if linha[0] != '\n':
                    if linha[entrada_saida[0]] == info:
                        # achamos o guerreiro pelo numero
                        if entrada_saida[1] == -1:
                            # So queremos verificar se ele existe em alguma base de dados
                            arquivo1 = arquivo1.replace(".txt", '')
                            arquivo1 = arquivo1.replace('Arquivos/CFG-', '')

                            return True, arquivo1
                        else:
                            return True, linha[entrada_saida[1]]

        return [False, 0]

    if arquivo is None:
        # Não recebemos nenhum arquivo, logo ou fazemos a busca inteligente, ou a burra.

        if info.isnumeric():
            # print('Estou entrando na recursao')
            # Podemos fazer a busca inteligente

            caso = busca_logica(info)
            # print(caso)

            if caso[0] is True:
                return caso[1]
            else:
                caso = str(int(caso) + 1)
                return pegando_info(info, entrada_saida, f'CFG-{caso}.txt', True)
        else:
            # print('Entrando na busca burra')
            caso = busca_burra()
            if caso[0]:
                return caso[1]
            else:
                return False

    else:
        # Estamos a receber algum arquivo.

        if count:
            # print("Estou fazendo a parte recursiva")
            # quer dizer que recebemos um subde busca logica

            try:
                # Vamos
                caso = busca_semiburra(arquivo)
                # print(caso)

                if caso[0]:
                    return caso[1]
                else:
                    # CFG-x.txt
                    for i in range(20, limite_de_busca):
                        if str(i) in arquivo:
                            caso = i + 1
                            break

                    return pegando_info(info, entrada_saida, f'Arquivos/CFG-{caso}.txt', count)

            except FileNotFoundError:
                # Caso esse arquivo não exista, não necessariamente quer dizer que já finalizou a busca
                # Vamos só fazer a busca burra de uma vez
                caso = busca_burra()
                if caso[0]:
                    return caso[1]
                else:
                    return False
        else:
            # print('Estou fazendo a busca semiburra')
            # quer dizer que só queremos fazer
            caso = busca_semiburra(arquivo)
            if caso[0]:
                return caso[1]
            else:
                return False


def verificador(chave, padrao):
    """Queremos identificar se duas palavras são quase iguais
    Não considera diferença entre maiusculo e minusculo, como verificador(pedro, PedRo) = 1"""
    i = 0

    try:
        while True:
            if chave[i].upper() != padrao[i].upper():
                return 0

            i += 1
    except IndexError:
        return 1


def torrando(num, puni, motivo, turm):
    # Aqui vamos precisar pensar bastante, pois há várias verificações que devem ser feitas, por exemplo,
    # devemos saber se a pessoa é a primeira a ser inserida no arquivo, sendo este necessariamente
    # criado, pois se ele já existe é porque em algum momento precisou-se para inserir alguem.
    # Deve ter o caso de ser a primeira torração do aluno ou o último caso, que apenas vamos adicionar
    # uma torração para a pessoa.
    # Sanha.

    # Sendo assim, só precisamos nos concentrar em tres casos, o de alguem ter a sua primeira torração ou
    # somente da adição de uma nova torração e o caso do arquivo estar a ser criado
    # 0 - Pessoa já presente, apenas nova torração
    # 1 - Arquivo já existe, mas a pessoa vai levar a sua primeira torração
    # 2 - Arquivo zerado! Não da para testar sem outra base de outro ano

    # Vamos verificar, criaremos essa variavel para podermos ter mais mobilidade de mudança de caso
    def verificando_se_presente():

        try:
            p = open(rf'Punicoes/punidos{turm}', 'r')
            linhas = p.readlines()

            # Devemos verificar se a referida pessoa esta presente
            for linha4 in linhas:
                if '\n' in linha4:
                    linha4 = linha4.replace('\n', '')

                if linha4.isnumeric():
                    if num == linha4:
                        # Achamos a pessoa no arquivo
                        return 0

            if len(linhas) == 0:
                # Pois está zerado
                p.close()
                return 2
            else:
                p.close()
                # Pois o cara so nunca havia sido torrado
                return 1

        except FileNotFoundError:
            # Se o arquivo não existe, precisamos criar e avisar
            q = open(rf'Punicoes/punidos{turm}', 'x')
            q.close()
            messagebox.showinfo(title='INFO', message='Base de informação de punidos criada')
            # Como acabou de ser criada, sabemos que esta COMPLETAMENTE vazia
            # Note que se o número não existe, nem chegar o código vai chegar
            return 2

    caso = verificando_se_presente()
    # print(f'Para o {num}, o caso é {caso}')
    # Aqui vamos criar a nossa lógica
    opcao = ['r', 'a', 'a']
    arquivo2 = open(rf'Punicoes/punidos{turm}', opcao[caso])

    if caso == 0:
        # Vamos ter que ler todas as linhas e dps vamos salvá-las com a modificação do sanha

        dados = arquivo2.readlines()

        arquivo2.close()

        # Agora devemos procurar o número que esta sendo torrado e inserir a nova torração no local certo
        achei = False
        for k in range(0, len(dados)):
            if dados[k].replace('\n', '') == num:
                # Quer dizer que achamos o guerreiro
                achei = True
                # print(f'Achei o aluno e estou a acrescentar a punição em {num}')
                # Para podermos sair do número, devemos ir uma unidade a mais

            if achei and dados[k] != '}\n':
                # print(f'Estou na linha {dados[k]} e vou adicionar a punição na posicao anterior')
                # quer dizer que chegamos ao final do referido aluno

                pinta = puni + '>' + motivo + '\n'
                dados.insert(k + 1, pinta)
                # print(f'Acabei de inserir o sanha nele')
                # Com esse insert, damos o shift para os lados
                break
                # Não vamos precisar varrer mais, por isso não há nada para colocarmos o }

        # agora vamos precisar escrever todos os dados
        arquivo2 = open(rf'Punicoes/punidos{turm}', 'w')

        for sanha in dados:
            arquivo2.write(sanha)

        arquivo2.close()
    elif caso == 1:
        # Logo, precisamos pular do outro {
        arquivo2.write(f'\n{num}\n{puni}>{motivo}\n{"}"}')
        arquivo2.close()
    elif caso == 2:
        # como vamos apenas adicionar uma pessoa zerada:
        arquivo2.write(f'{num}\n{puni}>{motivo}\n{"}"}')
        arquivo2.close()


def extrair_planilha(arquivo):
    import pandas as pd
    # Acho que é mais interessante fazermos pelo PDF, pois é o que conhecemos.
    # Em versões posteriores, pode-se adicionar essa função.

    def obtendo_turma(num):
        for arq in os.listdir('../SargenteanteAlpha/Arquivos'):
            if '.txt' in arq:
                if pegando_info(num, [0, -1], f'Arquivos/{arq}'):
                    # Achamos ele nesse arquivo
                    for i in range(20, limite_de_busca):
                        if str(i) in arq:
                            # Vamos retornar a turma
                            return i

    # Vamos abrir o sanha
    planilha = pd.read_excel(arquivo)

    # Vamos receber os números,
    # Descobrir a turma de cada um,
    # E adicionar a punição dele.

    # Note que nem vamos precisar do nome, vamos usar apenas os números.

    for index in range(0, len(planilha['Número'])):
        # Vamos obter o número,
        numero = str(planilha['Número'][index])

        # Vamos obter a punição dele
        punicao = str(planilha['Punição'][index])

        # Vamos obter o motivo
        moti = str(planilha['Fato Observado'][index]).upper()

        # Devemos fazer a verificação se há letras em acento
        diferentes = [['Ã', 'A'], ['Õ', 'O'], ['Á', 'A'], ['É', 'E'], ['Ó', 'O'], ['Ú', 'U'], ['Ç', 'C']]
        for diferen, normal in diferentes:
            if diferen in moti:
                moti = moti.replace(diferen, normal)

        # Vamos obter de qual turma ele é:
        turma = obtendo_turma(numero)

        # De posse da turma desse cara, vamos simplesmente adicionar o Sanha nele
        torrando(numero, punicao, moti, turma)


def extrair_PDF(arquivo):
    # Desistimos porque não conseguimos baixar o java de forma correta, vi tudo que poderei ser feito e não consegui,
    # perdão por tudo
    import fitz  # Importe a biblioteca PyMuPDF

    pdf_document = fitz.open(arquivo)

    gap = 0
    try:
        # Como não sabemos a quantidade
        while True:
            # vamos pegar a página
            pagina = pdf_document.load_page(gap)
            # print(f'Em teoria, estou pegando a página {gap}')

            # Pegando as informações da tabela que exista na página
            informacoes = pagina.get_text("text").split("\n")

            # Vamos receber sempre 2 tipos de arquivos, entretanto, em ambos há:
            # Mapa --> Número, Nome de Guerra, Seção, Ano, Fato Observado, Punição

            # Em teoria, há muitos espaços vazios, vamos agasalhar eles e verificar dps
            niveis_punicao = ['DRS1', 'DRS2', 'II', 'IA', 'FATD']
            diferentes = [['Ã', 'A'], ['Õ', 'O'], ['Á', 'A'], ['É', 'E'], ['Ó', 'O'],
                          ['Ú', 'U'], ['Ç', 'C'], ['À', 'A']]
            for index in range(0, len(informacoes)):
                numero = str(informacoes[index])
                if numero.isnumeric() and len(numero) > 2:
                    # É de fato o número de um aluno
                    # Devemos verificar de qual turma ele é
                    turma = pegando_info(numero, [0, -1])
                    if turma is not False:
                        # quer dizer que o sanhudo existe e temos a turma dele
                        # print(f'O {numero} existe na turma {turma}')

                        # Devemos pegar os motivos de forma correta, isto é, vamos avançar o Sanha das caixas até
                        # chegar numa caixa especifica e aí vamos parar de construir
                        j = 0
                        motivo = ''
                        while True:

                            texto = informacoes[index + 4 + j].strip().upper()
                            # print(f'Estou visualizando {texto}')
                            # Condição de finalização
                            caso_justificado = texto == 'JUSTIFICADO'
                            caso_finalizou = texto in niveis_punicao

                            if caso_justificado:
                                # Quer dizer que nem punido ele estará
                                # Logo, precisa sair
                                # print(f'Há justificação, logo vou anular')
                                break

                            if caso_finalizou:
                                # Se chegou ao final da punicao
                                punicao = texto
                                # print(f'Já sei qual é a punição dele e vou encerrar com {motivo}')
                                for diferen, normal in diferentes:
                                    if diferen in motivo:
                                        motivo = motivo.replace(diferen, normal)

                                # Vamos torrar ele e seguir para o proximo
                                # print(f'Vou torrar {numero} com {punicao} por causa de {motivo}')
                                torrando(numero, punicao, motivo, turma)

                                # Como fizemos tudo que precisavamos com esse aluno, vamos ao próximo
                                break

                            else:
                                # Se não chegou ao final da punicao, o que precisamos fazer é
                                motivo = motivo + texto

                            j += 1

            # print(f'Agora, vamos indo para a proxima pagina')

            # Para irmos para a próxima página, caso exista
            gap += 1
    except ValueError:
        # Feche o arquivo PDF
        pdf_document.close()

        # Já terminamos
        deci = messagebox.askyesno(message='Extração finalizada, deseja apagar o arquivo pdf?')

        if deci:
            os.remove(arquivo)
        else:
            pass

    except:
        return messagebox.showerror(message='Erro Sanhudo na Extração')


def criador_planilha_assistente(nome5):
    ano = int(datetime.now().year) - 2000

    def pegando_base():
        # Vamos precisar montar a base de informações
        total = {'REMAT': [],
                 'ATIVA': [],
                 'RESERVA': [],
                 }

        def pegando_classe(num):
            # Exemplo 23054
            num = str(num)
            # print(f'Estarei vendo {num[2]} e {num[:2]}')
            # print(f'Vou tentar comparar {int(num[:2])} e {int(ano)}')

            if num[2] == '4':
                # quer dizer que é reserva
                return 'RESERVA'

            elif int(num[:2]) == int(ano):
                return 'ATIVA'

            else:
                # remat com certeza
                return 'REMAT'

        with open(f'Arquivos/{nome5}', 'r') as pegando:
            for linha in pegando:
                linha = linha.split("-")

                if linha[0] != '\n':
                    if '\n' in linha:
                        linha = linha.replace("\n", '')

                    # Retirando o 0
                    linha.pop()
                    cls = pegando_classe(linha[0])

                    total[cls].append('-'.join(linha))
                    # print(f'Estou adicionando{linha} à classe {cls}')

        return total

    base_criada = pegando_base()

    def construindo_planilha_assistente(base_informacoes):
        # Vamos construir uma planilha Sanhuda capaz de safar o sgt

        sanha = xl.Workbook()

        pagina = sanha.active

        cores = {'ATIVA': PatternFill(start_color='0000FF00', end_color='0000FF00',
                                      fill_type='solid'),
                 'REMAT': PatternFill(start_color='00808080', end_color='00808080',
                                      fill_type='solid'),
                 'RESERVA': PatternFill(start_color='00FFFF00', end_color='00FFFF00',
                                        fill_type='solid')
                 }

        titulos = ['Número', 'Nome Completo', 'Nome de Guerra', 'Segmento']
        colunas_referenciadas = ['A', 'B', 'C', 'D']

        for i in range(0, 4):
            # Título da Coluna
            pagina[colunas_referenciadas[i] + '1'] = titulos[i]
            pagina[colunas_referenciadas[i] + '1'].font = Font(bold=True)

        # Vamos guardar a linha de cada um assim
        linha = 2
        # Vamos preencher com cada informação
        for chave in base_informacoes.keys():
            # Em cada chave, teremos uma lista de alunos a serem comentados

            # Vamos obter cada aluno de cada vez
            for aluno in base_informacoes[chave]:
                index = 0
                aluno = aluno.split('-')

                # Vamos adicionar as informações de cada aluno
                for info in aluno:
                    # Adicionamos
                    pagina[colunas_referenciadas[index] + str(linha)] = info
                    pagina.column_dimensions[colunas_referenciadas[index]].width = len(
                        info) + 5
                    pagina[colunas_referenciadas[index] + str(linha)].fill = cores[chave]

                    index += 1

                # Como terminamos com um aluno, devemos ir à outra linha
                linha += 1

            # Agora, todas as cores correspondentes

        sanha.save(nome5.replace(".txt", '') + '.xlsx')

    construindo_planilha_assistente(base_criada)


def preenchendo_servicos_gerais(lista):
    for arquivo in os.listdir('../SargenteanteAlpha/ServicosGerais'):
        houve_mudanca = False

        # Pegando as linhas completas
        p = open(f"ServicosGerais/{arquivo}", 'r')
        linhas = [linha.replace('\n', '') for linha in p.readlines()]
        p.close()

        # Procurando os sanhudos
        for numero, tipo in lista:
            index_numero = 0
            for linha in linhas:
                linha = linha.split('=')

                # Achamos o sanhudo
                if pegando_info(numero, [0, 1]) == linha[0]:
                    # Como vamos modificar
                    linhas[index_numero] = str(linhas[index_numero].split('=')[1])
                    linhas[index_numero] = linhas[index_numero].split(';')
                    # print(linhas[index_numero])

                    chegamos_ao_outro_lado = False
                    estamos_vendo_tipo = False
                    index = 0
                    for elemento in linhas[index_numero]:
                        elemento = elemento.split("-")
                        # print(elemento)

                        # Achamos o tipo procurado
                        # Vamos causar o Sanha
                        if elemento[0] == tipo:
                            linhas[index_numero][index] = linhas[index_numero][index].split('-')
                            # print(linhas[index_numero][index])

                            # Vamos mudar o Sanha
                            linhas[index_numero][index][1] = str(int(elemento[1]) + 1)

                            # Como já com certamos todx Sanha
                            # Vamos juntar tudo
                            linhas[index_numero][index] = '-'.join(linhas[index_numero][index])

                            # Terminamos
                            break

                        index += 1

                    linhas[index_numero] = ';'.join(linhas[index_numero])

                    linhas[index_numero] = '='.join([linha[0], linhas[index_numero]])

                    # Vamos finalizar com esse cara

                    # Já fizemos as modificações nesse cara
                    houve_mudanca = True

                    break

                index_numero += 1

        if houve_mudanca:
            # Vamos salvar as mudanças
            os.remove(f"ServicosGerais/{arquivo}")

            primeiro = True
            with open(f"ServicosGerais/{arquivo}", 'x') as sg:
                for linha in linhas:
                    if primeiro:
                        sg.write(linha)
                        primeiro = False
                    else:
                        sg.write(f'\n{linha}')

        # Finalizei o Sanha


def gerar_treeview(base):
    # Vamos ler o arquivo de serviços e montar um treeview bolado para nós vermos
    # Vamos pegar o Sanha e varrer tudo

    if messagebox.askyesnocancel(message=
                                 'Você está prestes a criar uma janela que mostrará quantos serviços cada um tirou em cada respectivo tipo de serviço.'):
        pass
    else:
        return 0

    turma = 0
    for i in range(20, limite_de_busca):
        if str(i) in base:
            turma = i
            break

    p = open(f"ServicosGerais/servicos{turma}", 'r')
    linhas = [linha.replace("\n", '') for linha in p.readlines()]
    p.close()

    # De posse das linhas
    # Vamos ter que produzir o Sanha de cada um.
    def organizando_sanha(linhas_totais):
        resultado = []
        index_referenciacao = 0
        for index in range(0, len(linhas_totais)):
            # Vamos cortar cada um
            linhas_totais[index] = linhas[index].split("=")
            linhas_totais[index][0] = pegando_info(linhas_totais[index][0], [1, 2])

            resultado.append([])
            resultado[index_referenciacao].append(linhas_totais[index][0])

            # Vamos trabalhar na parte dos números agora
            for elemento in linhas[index][1].split(";"):
                tipo, quant = elemento.split('-')
                resultado[index_referenciacao].append(quant)

            index_referenciacao += 1

        return resultado

    base = organizando_sanha(linhas)

    pesos = [p for p in range(90, 0, -10)]

    def ordenando(lista):
        # Vamos comparar seguindo a partir do primeiro tipo de serviço e sua respectiva quantidade

        # Vamos providenciar a solução do aluno NEVES
        resultado = 0
        for w in range(1, len(lista)):
            resultado += int(lista[w]) * pesos[w - 1]

        return resultado

    base.sort(key=lambda sublista: ordenando(sublista), reverse=True)

    def apresentacao():

        def colocando_treeview(mestre):
            tipos = {
                "1": 'PtSegMasc',
                "1.2": 'PtSegFem',
                "1.4": '3° Piso',
                "1.8": 'Cabo',
                "2": 'Sentinela',
                "3": 'Cabo da Guarda',
                "3.5": 'Cdt da Guarda',
                "4": 'Sgt',
                "5": 'Aux Of'
            }
            colunas = ['Nome de Guerra']
            for tipo in tipos.values():
                colunas.append(tipo)

            tv = ttk.Treeview(mestre, columns=colunas, show='headings')

            tamanho = [90, 50]
            i1 = 0
            for coluna in colunas:
                tv.heading(coluna, anchor='center', text=coluna)
                tv.column(coluna, minwidth=20, width=tamanho[i1], anchor='center')
                if i1 == 0:
                    i1 += 1

            tv.place(x=10, y=10, width=550, height=300)

            for sublista in base:
                tv.insert("", 'end', values=sublista)

        janela = criando_janela('Serviços Gerais', 570, 400)

        colocando_treeview(janela)

        # Disponibizando a forma de procura de alguém específica
        def disponbilizando_filtro():
            # Colocando entrada
            Label(janela, text='Insira o nome:', bg='#dde').place(x=10, y=320)
            nome_ = Entry(janela)
            nome_.place(x=10, y=350, width=90, height=20)
            Label(janela, text='Press Enter', bg='#dde').place(x=10, y=370)

            def filtrando(event):
                # Vamos procurar e pegar

                nome = nome_.get()
                for elemento in base:
                    if verificador(elemento[0], nome):
                        # Achamos o Sanha, devemos apresentar as informações
                        Label(janela, text=elemento[0], bg='#dde').place(x=100, y=320)
                        for q in range(1, 10):
                            Label(janela, text=elemento[q], bg='#dde').place(x=75 + q * 50, y=350)

                        # Se achamos, vamos apenas limpar o espaço
                        return nome_.delete(0, 'end')

                # Se não acharmos
                return messagebox.showwarning(message='Não encontrado')

            nome_.bind('<Return>', filtrando)

        disponbilizando_filtro()

        janela.mainloop()

    apresentacao()
