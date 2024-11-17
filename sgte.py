from tkinter import messagebox
from tkinter import *
from tkinter import ttk
from datetime import datetime  # Para conseguirmos mexer com tempo muito fácil
import os  # para conseguirmos visualizar as listas das pastas

# Vamos pegar as subfunções que serão uteis
from subfuncoes import codigo_mes, limpeza, obtendo_hoje, serv, criando_janela, pegando_info, verificador, \
    seguinte, informador, historicando, enviando_email, enviando_sms, obtendo_data, destruidor, \
    extrair_planilha, extrair_PDF, torrando, criador_planilha_assistente, preenchendo_servicos_gerais, gerar_treeview

from criador_do_sanha import escalador

# Número de Reestruturações complexas: 4
# Dia de início: 30 de julho

"""Seguinte, vamos fazer o seguinte:
        # Já terminamos tudo, em teoria está tudo construído
        
        -- Problemas Relatados:
        -Colocar Destruidor em toda janela secundária
        -***Verificar toda a forma de aplicação de escalas == Em teoria, está dando certo, cuidado com isso
        -Lembre-se de atualizar a documentação
        
        Vamos ter 3 versões sanhudas:
        
        --- Alpha: por meio do código fonte, vamos fazer todas as verificações possíveis e colocar diversos try/except
                   pelo código, resolvendo os sanhas que aparecerem e, além disso, refinando o código além de documentá-lo.
                   Vamos tentar também deixar o código mais intuitivo para o usuário.
                   Em teoria, vai ser a etapa mais demorada.
                   Lembra de colocar um 'By MathDeyvi'.
        
        
        --- Beta:  Vamos transformar o conjunto inteiro de sanhas em um único executável e bradá-lo. Vamos ter um arquivo
                   .exe e que vai rodar tudo perfeitamente, em teoria, por isso faremos mais análises, mas será bem mais
                   rápido que a etapa Alpha.
        
        
        --- Gamma: Vamos pegar o .exe e dá-lo para alguém, ou o sgte do primeiro ou do segundo ano. Ele usará o programa
                   relatando os sanhas que obter e as coisas que não entender muito bem ou possíveis erros que só alguém
                   fora da sua cabeça mathdeyviniana poderia pensar.
                   É uma ótima hora de conseguir ideias e sugestões.
                   
        
        --- Oficial: Por fim, vamos pedir sugestões aos que conhecem o fato e tentar apresentar o projeto final.
    
        
        
        Reestruturação
        {
            Salvar os numeros de cada aluno
            
            Mudar forma de datar as escalas
            
            Reconstruir o sistema do criador do sanha para introduzir a situação de cada turma aos seus correspondentes
            serviços. - sanha
            
            MUDAR TODA A FORMA DE BUSCA e de salvamento para a forma numerica dos aluno - sanha infinito
            
        }
        
        terminou 25/10/23
        
        """

versao = 'v1.4'


def sgte():
    # Criando variaveis que serão necessárias
    ANO_HOJE = datetime.now().year
    DATA_HOJE = obtendo_hoje()

    # Devemos ter uma função que vai cuidar das escalas e da forma de armazenamento
    # Por exemplo, no fim do mes todas elas devem ir para um armazenamento
    # ACHO MELHOR COLOCAR TODx MEIO DO ME

    janela = criando_janela("SargenteanteAlpha", 600, 400)

    note1 = ttk.Notebook(janela)
    note1.place(x=0, y=0, width=600, height=400)

    # Agora criando o espaço que colocaremos as informacoes
    def criando_ambientes(mestre):
        ambi = []
        classes = ['Menu de Serviço', 'Menu de Baixados', 'Menu de Punidos']
        for clas in classes:
            p = Frame(mestre, bg="#dde")
            mestre.add(p, text=clas)
            ambi.append(p)
        return ambi

    escala_servico, menu_baixados, menu_punidos = criando_ambientes(note1)

    def colocando_autor():
        papeis = [escala_servico, menu_baixados, menu_punidos]
        for papel in papeis:
            Label(papel, text=versao, bg='#dde').place(x=480, y=350)
            Label(papel, text='By MathDeyvi', bg='#dde').place(x=510, y=350)

    colocando_autor()

    # Vamos criar o espaço para fazer os menus de funções especiais
    def criando_menu():
        # Criando a barra de menu
        barra_menu_ = Menu(janela)
        barra_menu1 = Menu(barra_menu_)
        return barra_menu1

    barra_menu = criando_menu()

    # Onde as informações serão salvas
    global alunos, limite_de_busca
    limite_de_busca = 50  # Até que número o serviço irá buscar
    alunos = 0
    informacoes = {'Nome Completo': [], 'Nome de Guerra': [], 'Seg': [], 'Serv': [], 'ID': []}

    def mexendo_servicos():
        """Aqui, nos aprisionaremos ao funcionamento da escala de serviço."""

        def criando_inicio():
            try:
                # Criando as molduras
                bele = Frame(escala_servico, borderwidth=2, relief="solid", bg="#dde")
                bele.place(x=0, y=0, width=370, height=370)
                possivel_base = []

                # Quais são as possíveis bases disponíveis
                try:
                    for possivel in os.listdir('../SargenteanteAlpha/Arquivos'):
                        if '.txt' in possivel:
                            possivel_base.append(possivel)
                except FileNotFoundError:
                    messagebox.showwarning(message='Ou você não inicializou corretamente ou tirou da pasta correta.')

                # Precisamos analisar as bases possiveis e mostrá-las ao usuario
                bases_possiveis1 = ttk.Combobox(escala_servico, values=possivel_base)
                bases_possiveis1['state'] = 'readonly'
                bases_possiveis1.set(possivel_base[0])
                bases_possiveis1.place(x=10, y=10)

                return bases_possiveis1
            except IndexError:
                messagebox.showerror(message='Não há nenhuma base de dados de alunos para serem mostrados.')
                return None

        bases_possiveis = criando_inicio()

        def visualizando_listageral():
            """Aqui, vamos organizar todas as informações sobre os alunos para mostrar.
            Teremos várias planilhas em excell para serem mostradas.
            Toda a nossa base de dados vai se basear primmeiramente na planilha em excell.
            """

            try:
                # Criando o botão para mostrar o treeview dos alunos com todas as informações
                def criando_treeview():

                    # Função para extrair e obter as informações gerais
                    def extraindo_geral():
                        # Definindo de que base extrairemos
                        base = bases_possiveis.get()

                        # Devemos limpar as listas para não ficar só aumentando de tamanho, pode dar erro facilmente
                        def zerando():
                            global alunos
                            alunos = 0
                            informacoes['ID'].clear()
                            informacoes['Nome Completo'].clear()
                            informacoes['Nome de Guerra'].clear()
                            informacoes['Seg'].clear()
                            informacoes['Serv'].clear()

                        zerando()

                        # Extraindo
                        with open(rf'Arquivos/{base}', "r") as arq:
                            global alunos
                            for linha in arq:
                                # lendo a linha
                                info = linha.split('-')

                                # Para que obviamente pule a linha que não contenha nada
                                if info[0] != '\n':
                                    alunos += 1
                                    informacoes['ID'].append(int(info[0]))
                                    informacoes['Nome Completo'].append(info[1])
                                    informacoes['Nome de Guerra'].append(info[2])
                                    informacoes['Seg'].append(info[3])
                                    informacoes['Serv'].append(int(info[4]))

                    extraindo_geral()

                    # Função para colocar o treeview com as informações na tela
                    def mostrando_treeview():

                        # Criando o TreeView
                        tv = ttk.Treeview(escala_servico,
                                          columns=['ID', 'Nome Completo', 'Nome de Guerra', 'Seg', 'Serv'],
                                          show='headings')

                        # Colocando as colunas do treeview
                        tv.column('Nome Completo', minwidth=0, width=150)
                        tv.heading('Nome Completo', text='Nome Completo')

                        tv.column('Nome de Guerra', minwidth=0, width=90)
                        tv.heading('Nome de Guerra', text='Nome de Guerra')

                        tv.column('Seg', minwidth=0, width=30)
                        tv.heading('Seg', text='Seg')

                        tv.column('Serv', minwidth=0, width=30)
                        tv.heading('Serv', text='Serv')

                        tv.column('ID', minwidth=0, width=40)
                        tv.heading('ID', text='ID')

                        tv.place(x=10, y=80, height=280)

                        # Aplicando as informações do TreeView
                        for c in range(0, alunos):
                            # print(informacoes['Nome Completo'][c])
                            tv.insert("", "end", values=[informacoes['ID'][c],
                                                         informacoes['Nome Completo'][c],
                                                         informacoes['Nome de Guerra'][c],
                                                         informacoes['Seg'][c],
                                                         informacoes['Serv'][c]])

                    mostrando_treeview()

                Button(escala_servico, text='Visualizar', command=criando_treeview).place(x=160, y=10, height=20)

                def gerando():
                    if messagebox.askyesnocancel(
                            message='Você está prestes a criar uma planilha assistente deste ano.'):
                        # Precisamos descobrir o nome da turma que está sendo mostrada
                        criador_planilha_assistente(bases_possiveis.get())

                Button(escala_servico, text='Planilha Assistente', command=gerando).place(x=250, y=10, height=20)

                Button(escala_servico, text='Visualização de Serviços',
                       command=lambda: gerar_treeview(bases_possiveis.get())).place(x=222, y=40,
                                                                                    height=20)

                # Disponibilizando os filtros possiveis
                filtros_possiveis = ['Ordem Alfabetica',
                                     'Serviços',
                                     'Segmento',
                                     'Numero']
                filtros = ttk.Combobox(escala_servico, values=filtros_possiveis)
                filtros['state'] = 'readonly'
                filtros.set(filtros_possiveis[0])
                filtros.place(x=10, y=40)

                # Aqui temos um sistema de filtros, para que assim possamos ter acesso às informações que desejamos mais rapidamente
                def aplicando_filtro():
                    """Aqui, o segredo é ficar a criar treeview em cima de outras!"""

                    if bases_possiveis is None:
                        messagebox.showerror(message='Não há nenhuma selecionada para filtrar')
                        return 0

                    tipo = filtros.get()

                    # Note que nenhuma informação vai ser alterada, nós só queremos organizá-las para priorizar algumas
                    # Sendo assim, vamos guardar o índice de cada aluno nessa ordem desejada
                    ordem = []

                    def filtrando(lista):

                        # Criando o TreeView
                        tv = ttk.Treeview(escala_servico,
                                          columns=['ID', 'Nome Completo', 'Nome de Guerra', 'Seg', 'Serv'],
                                          show='headings')

                        # Colocando as colunas do treeview
                        tv.column('Nome Completo', minwidth=0, width=150)
                        tv.heading('Nome Completo', text='Nome Completo')

                        tv.column('Nome de Guerra', minwidth=0, width=90)
                        tv.heading('Nome de Guerra', text='Nome de Guerra')

                        tv.column('Seg', minwidth=0, width=30)
                        tv.heading('Seg', text='Seg')

                        tv.column('Serv', minwidth=0, width=30)
                        tv.heading('Serv', text='Serv')

                        tv.column('ID', minwidth=0, width=40)
                        tv.heading('ID', text='ID')

                        tv.place(x=10, y=80, height=280)

                        # Aplicando as informações do TreeView
                        for indice in lista:
                            # print(informacoes['Nome Completo'][c])
                            tv.insert("", "end", values=[informacoes['ID'][indice],
                                                         informacoes['Nome Completo'][indice],
                                                         informacoes['Nome de Guerra'][indice],
                                                         informacoes['Seg'][indice],
                                                         informacoes['Serv'][indice]])

                    if tipo == 'Ordem Alfabetica':

                        quant_nome = []
                        k = 0
                        for nome_completo in informacoes['Nome Completo']:
                            quant_nome.append([k, nome_completo])
                            k += 1

                        quant_nome.sort(key=lambda sublista: sublista[1])

                        for id2, nome in quant_nome:
                            ordem.append(id2)

                        filtrando(ordem)

                    elif tipo == 'Serviços':

                        # Para sabermos quanto de serviço cada um já executou
                        quant_serv = []
                        k = 0
                        for serv1 in informacoes['Serv']:
                            # Afinal precisamos à quem aquela quantidade de serviço se refere.
                            quant_serv.append([k, serv1])

                            k += 1

                        # Agora vamos ordenar essas sublistas baseando-se no elemento de serviço
                        quant_serv.sort(key=lambda sublista: sublista[1], reverse=True)
                        # O key é para conseguirmos usar um novo método de ordenação

                        # Erro Que Tinhamos
                        """
                        Estava dando um erro sério, pois conforme se colocava para visualizar a lista geral sem filtro e dps se
                        aplicava o filtro, acontecia isso:
                        [[1, 5], [0, 0], ...
                        [[1, 5], [33, 5], [0, 0],...
                        [[1, 5], [33, 5], [65, 5],...

                        ~O ERRO VEM DESSA NOVA FORMA DE FAZER O SORT~
                        Equivocado, após análises, pudemos conferir que o erro vem de como salvamos as informações em Serv
                        Note que o metodo sort não pode estar errado. 
                        E que tanto ordem quanto quant_serv são zerados aqui, então so pode vir de informações['serv']

                        O erro era porque não zeravamos a lista


                        """

                        # Sendo assim, basta salvarmos a ordem certa
                        for id1, qua in quant_serv:
                            ordem.append(id1)

                        filtrando(ordem)

                    elif tipo == 'Segmento':

                        # Vamos pegar so o seg.fem porque elas são muito minoria
                        for p in range(0, alunos):
                            if informacoes['Seg'][p] == 'F':
                                ordem.append(p)

                        filtrando(ordem)

                    elif tipo == 'Numero':

                        quant_num = []
                        k = 0
                        for n1 in informacoes['ID']:
                            quant_num.append([k, n1])
                            k += 1

                        quant_num.sort(key=lambda sublista: sublista[1])

                        for nu, num in quant_num:
                            ordem.append(nu)

                        filtrando(ordem)

                    """Antes tínhamos um código criando o treeview em cada if, agora está mais inteligente"""

                Button(escala_servico, text='Filtrar', command=aplicando_filtro).place(x=160, y=40, height=20)

                Label(escala_servico,
                      text='Clique em Visualizar para ver alguma lista de alunos e sinta-se \n livre para ordená-la',
                      bg='#dde').place(x=20, y=180)
            except AttributeError:
                messagebox.showerror(message='Não há base de alunos disponível')

        visualizando_listageral()

        def vendo_descanso():
            """
            Aqui, aprisionaremos quem estará na lista de descanso, para não entrar no serviço.

            O padrão que vamos fazer é ter:   ID-DIAQUESAIDODESCANSO

            Um erro que voce pode pensar é: Pq não mostra todos os nomes de cada um. Entretanto, esta função mostra-
            rá apenas os militares que pertencem à turma que esta sendo visualizado
            """

            # Criando moldura
            descanso_livre = LabelFrame(escala_servico, borderwidth=2, relief="solid", bg="#dde",
                                        text="Lista de Descanso")
            descanso_livre.place(x=380, y=10, width=210, height=80)

            # Função que serve para ver quem está no periodo do descanso
            def ver_descanso():

                # Criando uma forma de guardar os interessados
                lista_descanso = []

                # lista_descanso[x][0] -- ID do aluno
                # lista_descanso[x][1] -- DIA QUE SAI DO DESCANSO

                # Função para extrair as informações dos descansados
                def extraindo_descanso():
                    """"""
                    """Aqui vamos fazer uma forca. 
                    Se não fosse os remats, poderiamos apenas colocar apenas o numero lido mais 4 e ver se o resultado 
                    consta na base, entretanto, não é nossa realidade.

                    Em outra situação ainda muito boa, poderiamos colocar um sistema de busca binária, mas note que o 
                    número dos remats não segue uma ordem sucessiva como os outros, ou seja, só nos resta o sanha."""

                    try:
                        with open(r'../SargenteanteAlpha/SubArquivos/Descansos') as desc:
                            for linha in desc:
                                linha_ = linha.split('-')
                                lista_descanso.append([linha_[0], int(linha_[1])])

                    except FileNotFoundError:
                        messagebox.showerror(title='ERROR', message='Erro ao abrir dados de Descanso')

                extraindo_descanso()

                if len(lista_descanso) == 0:
                    return messagebox.showwarning(message='Não há ninguém no período de descanso.')

                # Vamos ordenar decentemente
                lista_descanso.sort(key=lambda sublista: sublista[1])

                # Devemos agora pensar em como mostrar essa lista de descanso de forma apropriada.
                def mostrando():
                    sanha = Tk()
                    try:

                        def colocando_treeview(mestre):
                            mestre.title("Escala de Descanso")
                            mestre.geometry("300x200")
                            mestre.configure(bg="#dde")
                            mestre.resizable(False, False)

                            mito = Frame(mestre, borderwidth=2, relief='solid', bg="#dde")
                            mito.place(x=10, y=10, width=280, height=180)

                            tv = ttk.Treeview(mito, columns=['Nome de Guerra', 'FINAL'],
                                              show='headings')

                            tv.column('Nome de Guerra', minwidth=0, width=60, anchor='center')
                            tv.heading('Nome de Guerra', text='Nome de Guerra')
                            tv.column('FINAL', minwidth=0, width=30, anchor='center')
                            tv.heading('FINAL', text='FINAL')
                            tv.place(x=10, y=10, height=155, width=255)

                            for id3, final in lista_descanso:
                                nome_guerra = pegando_info(id3, [0, 2])
                                if nome_guerra is False:
                                    messagebox.showwarning(
                                        message='Há uma pessoa que não está em nenhuma base de dados.')

                                tv.insert("", "end", values=[nome_guerra,
                                                             final])

                            """Estava dando um erro, pois estavamos usando o ID como indice da lista."""

                        colocando_treeview(sanha)

                        sanha.mainloop()

                    except IndexError:

                        messagebox.showerror(title='ERROR', message='Erro ao tentar visualizar lista de descanso')

                mostrando()

            Button(descanso_livre, text='Ver Descansos', command=ver_descanso).place(x=65, y=20)

            informador(descanso_livre, 160, 20,
                       """Mostra uma lista dos alunos que estão sob o tempo de descanso, isto é, tiraram serviço recentemente.""")

            # Função que serve para verificar quem está a sair do descanso
            def verificando_descansados():
                """"""
                """Aqui vamos apenas verificar se alguem que esta com final registrado dia x e se é dia x.
                Ao criarmos as escalas, vamos registrar sempre x + 3, onde x é o dia inicial de serviço."""

                # Vamos guardar as informações aqui
                todos = []

                # todos[x][0] -- ID
                # todos[x][1] -- FIM
                # Vamos pegar as informações
                def obtendo_todos():
                    """Vamos abrir e guardar todos, para que só assim possamos escrever quem não acabou"""

                    with open(r'../SargenteanteAlpha/SubArquivos/Descansos') as desc:
                        for linha_ in desc:
                            linha = linha_.split('-')
                            todos.append([linha[0], int(linha[1])])

                obtendo_todos()

                # Agora, vamos reescrever o arquivo lembrando de não colocar quais chegaram ao fim
                def reescrevendo():

                    # Vamos apagá-lo e criá-lo de novo
                    os.remove(r'../SargenteanteAlpha/SubArquivos/Descansos')
                    arq = open(r'../SargenteanteAlpha/SubArquivos/Descansos', 'x')

                    k = True
                    q = 0  # quantidade para avisarmos o usuário
                    for ID, fim in todos:
                        if fim != int(DATA_HOJE.replace(DATA_HOJE[-1], '')):
                            if k:
                                arq.write(f'{ID}-{fim}')
                                k = False
                            else:
                                arq.write(f'\n{ID}-{fim}')
                        else:
                            q += 1

                    if q != 0:
                        messagebox.showinfo(title='DESCANSADOS', message=f'Um total de {q} sairam do descanso')

                    arq.close()

                reescrevendo()

            # Essa função deve funcionar de maneira autonoma
            verificando_descansados()

        vendo_descanso()

        def vendo_escala():
            """Aqui vamos definir como será o formato das nossas escalas.
            ID-TIPODESERVIÇO"""

            # Devemos mostrar as escalas que existem e deixar que o usuario escolha qual deseja ver.
            # Por falar nisso, vamos mostrar a escala no Sanha ou mostrar em pdf?

            Lista_Escalas = LabelFrame(escala_servico, text='Lista de Escalas', bg="#dde", relief='solid')
            Lista_Escalas.place(x=380, y=130, width=210, height=80)

            def observador():
                # Aqui, devemos criar uma janela e mostrar um combobox das escalas disponiveis

                def criando_ambiente():

                    mostrando1 = criando_janela("Exibindo Escalas Disponiveis", 300, 200)

                    mito1 = Frame(mostrando1, borderwidth=2, relief='solid', bg="#dde")
                    mito1.place(x=10, y=10, width=280, height=180)

                    # Vamos pegar as escalas que podem ser vistas
                    escala_possiveis = []
                    for escala_possivel in os.listdir(r'../SargenteanteAlpha/Escalas'):
                        escala_possiveis.append(escala_possivel)

                    if len(escala_possiveis) == 0:
                        messagebox.showwarning(message='Não há escalas para ver')
                        mostrando1.destroy()
                        return [None, None, None]

                    escalas_mostradas = ttk.Combobox(mito1, values=escala_possiveis)
                    escalas_mostradas['state'] = 'readonly'
                    escalas_mostradas.set(escala_possiveis[0])
                    escalas_mostradas.place(x=70, y=10)

                    return [mito1, mostrando1, escalas_mostradas]

                # Uma função retornar uma lista é ótimo pqp
                mito, Mostrando, escala_ = criando_ambiente()

                if Mostrando is None:
                    return 0

                # Vamos colocar varios radionbutton para ver como o usario desejará ver
                opcao = IntVar(mito)
                opcao.set(3)

                Radiobutton(mito, text='SMS', variable=opcao, value=1, bg="#dde").place(x=5, y=40)
                Radiobutton(mito, text='EMAIL', variable=opcao, value=2, bg="#dde").place(x=110, y=40)
                Radiobutton(mito, text='Janela', variable=opcao, value=3, bg='#dde').place(x=210, y=40)

                def visualizador():

                    # Vamos obter a opção do usário
                    opcao_ = opcao.get()

                    # Vamos pegar a escala desejada
                    escala = escala_.get()

                    Mostrando.destroy()

                    # Independente da forma como visualizaremos, precisaremos de uma função capaz de identificar o tipo de
                    # serviço de cada militar
                    # ESCREVEMOS A FUNÇÃO DE SERVIÇO NO SUBFUNÇÕES

                    # Independente de como vamos visualizar, também precisamos das informações
                    guarnicao = []

                    # Função para extrair e preencher guarnicao
                    def extraindo_escala():
                        """"""
                        """Aqui, podemos ter um serio problema.
                        Imagine que vamos ter varias pessoas de varios anos aqui, pode ser que não funcione devidamente.
                        Sendo assim, vamos guardar os numeros de cada militar e partir para o abraço. 

                        Aqui, só estamos interessados em pegar os militares do ano? Não, essa deve ser uma função capaz de
                        mostrar todos os dados...
                        Talvez não devessemos nem colocar por numero e sim por nome!!

                        """

                        with open(rf'Escalas/{escala}') as esc:
                            for linha_ in esc:
                                linha = linha_.split("-")

                                guarnicao.append([linha[0], serv(linha[1])])

                    # Preenchendo
                    extraindo_escala()

                    def gerando_mensagem(escala1, guarnicao1):
                        texto = f'\n{obtendo_data(escala1)}\n\n'

                        k = True
                        anterior = ''
                        for idx, tipo in guarnicao1:
                            if anterior != tipo:
                                texto = texto + '\n'
                                anterior = tipo
                            nome = pegando_info(idx, [0, 2])
                            if nome is False:
                                messagebox.showwarning(
                                    message='Há uma pessoa que não está em nenhuma base de alunos, cuidado.')

                            if k:
                                k = False
                                texto = texto + f'{nome}' + 1 * ' = ' + f'{tipo}'
                            else:
                                texto = texto + f'\n{nome}' + 1 * ' = ' + f'{tipo}'

                        return texto

                    if opcao_ == 1:
                        # quer dizer que queremos ver a escala num SMS para alguem

                        def enviando():

                            # Devemos saber para que número sera o envio do SMS
                            global numero
                            numero = 0

                            def perguntando():
                                jan2 = Tk()
                                jan2.title(escala)
                                jan2.configure(bg="#dde")
                                jan2.geometry("250x150")
                                jan2.resizable(False, False)

                                p = Frame(jan2, bg='#dde', relief='solid', borderwidth=1)
                                p.place(x=10, y=10, height=150 - 20, width=230)

                                Label(p, text='Insira ou escolha o número a ser enviado:', bg='#dde').place(x=5, y=10)

                                def pegando1(evento):
                                    global numero
                                    numero = num.get().replace("\n", '')

                                    if numero.isnumeric():
                                        # Então conseguimos

                                        if numero not in numeros:
                                            if messagebox.askyesno(title='Salvar',
                                                                   message='Deseja salvar esse contato na lista de históricos?'):
                                                historicando(numero, True)

                                        enviando_sms(gerando_mensagem(escala, guarnicao), numero)

                                        jan2.destroy()

                                    else:
                                        # Deu ruim na entrada
                                        messagebox.showerror(title='ERROR', message='Houve um erro no número colocado')
                                        jan2.destroy()

                                numeros = historicando(1, False)
                                num = ttk.Combobox(jan2, values=numeros)
                                num.set(numeros[0])
                                num.place(x=15, y=40)
                                Label(jan2, text='Press Enter', bg='#dde').place(x=15, y=70)
                                num.bind('<Return>', pegando1)

                                jan2.mainloop()

                            # Salvando o numero
                            perguntando()

                            # Agora, devemos enviar a mensagem da escala

                        enviando()

                    elif opcao_ == 2:
                        # Vamos enviar por endereço eletrônico, em forma de prompt.

                        # Devemos obter o email do destinatário
                        jan = criando_janela('Pegando Destinatarios', 300, 150)

                        def mudar_credenciais():
                            # vamos deixar uma maneira de mudar as credenciais do usuário.
                            jan.destroy()

                            jan1 = criando_janela('Mudando credenciais', 300, 200)

                            Label(jan1, text='Insira o email de quem vai enviar:', bg='#dde').place(x=10, y=10)
                            remet = Entry(jan1)
                            remet.place(x=10, y=40)

                            Label(jan1, text='Insira a senha de aplicativo criada:', bg='#dde').place(x=10, y=70)
                            senha1 = Entry(jan1)
                            senha1.place(x=10, y=100)

                            def salvar():
                                remetente = remet.get()
                                senha = senha1.get()

                                if '.com' in remetente and '@' in remetente:
                                    # De posse do Sanha.

                                    p = open('../SargenteanteAlpha/SubArquivos/HistoricoVisual')
                                    linhas = p.readlines()
                                    p.close()

                                    linhas[-1] = f'{remetente};{senha}'

                                    os.remove('../SargenteanteAlpha/SubArquivos/HistoricoVisual')

                                    with open('../SargenteanteAlpha/SubArquivos/HistoricoVisual', 'x') as hv:
                                        for linha in linhas:
                                            hv.write(linha)

                                    messagebox.showinfo(message='Sucesso na mudança de credenciais')
                                    jan1.destroy()
                                    jan.destroy()
                                else:
                                    return messagebox.showerror(message='Email não válido')

                            Button(jan1, text='Salvar', command=salvar).place(x=10, y=130)

                            jan1.mainloop()

                        Button(jan, text='Mudar credenciais', command=mudar_credenciais).place(x=10, y=80)

                        def pegando(event):
                            email = combo.get()
                            email = email.replace("\n", '')

                            if email not in opcoes:
                                if messagebox.askyesno(message='Deseja salvar esse contato?'):
                                    historicando(email, True)

                            texto = gerando_mensagem(escala, guarnicao)

                            enviando_email(email, texto)

                            jan.destroy()

                        Label(jan, text='Insira ou escolha email de quem vai receber a escala', bg='#dde').place(x=10,
                                                                                                                 y=20)
                        opcoes = historicando(2, False)
                        combo = ttk.Combobox(jan, values=opcoes)
                        Label(jan, text='Press Enter', bg='#dde').place(x=220, y=50)
                        combo.place(x=10, y=50, width=200)
                        combo.set(opcoes[0])
                        combo.bind('<Return>', pegando)

                        jan.mainloop()

                    elif opcao_ == 3:
                        # Tivemos que abandonar a ideia da escala por PDF pois ela seria totalmente inviável.

                        def janela_escala(pessoal):

                            jan3 = Tk()
                            jan3.title(escala)
                            jan3.configure(bg="#dde")
                            # Vamos precisar inserir a altura total com base na quantidade de militares na guarnição
                            altura = 20 + 40 * int(len(guarnicao))
                            jan3.geometry(f"250x{altura}")
                            jan3.resizable(False, False)

                            p = Frame(jan3, bg='#dde', relief='solid', borderwidth=1)
                            p.place(x=10, y=10, height=altura - 20, width=230)

                            k = 0  # Para teros variavel de controle
                            # vamos criar um loop e ver como o Sanha se comporta
                            for idx, tipo in pessoal:
                                # Vamos colocar um label com o outro em cima! Genial

                                Label(p, text=f'{pegando_info(idx, [0, 2])}' + ' - ' * 5 + tipo, bg='#dde').place(x=10,
                                                                                                                  y=10 + k * 30)
                                k += 1

                            jan3.mainloop()

                        janela_escala(guarnicao)

                Button(mito, text='Visualizar', command=visualizador).place(x=100, y=140)

                def apagador_de_escala():
                    escala = escala_.get()

                    try:
                        os.remove(f'Escalas/{escala}')

                        Mostrando.destroy()

                        observador()
                    except:
                        return messagebox.showerror(message='Erro ao tentar apagar a escala')

                Button(mito, text='Apagar Escalar', command=apagador_de_escala).place(x=90, y=90)

                Mostrando.mainloop()

            Button(Lista_Escalas, text='Ver Escalas', command=observador).place(x=72, y=20)

            informador(Lista_Escalas, 160, 20,
                       'Mostra a lista de escalas que poderão ser visualizadas de 3 formas diferentes')

            def sanha():
                messagebox.showinfo(title='Registro',
                                    message='Na pasta da lixeira, no arquivo .xlsx está o histórico de todas as escalas já efetivadas.')

            Button(Lista_Escalas, text=' + ', command=sanha).place(x=20, y=20)

        vendo_escala()

        def criando_escala():
            """Aqui vamos desenvolver toda a nossa teoria de criação de escalas,
            para então criarmos a função de fato em outro local

            Para criarmos uma escala, vamos criar uma lista formada de outras listas sendo compostas pela ID do aluno e
            também pela quant. Serviços Também devemos ter noção se há segfem nela"""

            Criar_Escala = LabelFrame(escala_servico, text='Ferramentas de Escalas', bg="#dde", relief='solid')
            Criar_Escala.place(x=380, y=250, width=210, height=80)

            def criador_do_sanha():
                # Devido à dificuldade excessiva do problema, devemos fazer isso em outro lugar!
                escalador()

            Button(Criar_Escala, text='Manipulação de Escalas', command=criador_do_sanha).place(x=22, y=20)

            informador(Criar_Escala, 165, 20,
                       'Disponibiliza todas as ferramentas possíveis para a criação e manutenção das escalas')

        criando_escala()

        def aplicando_escala():
            """
            """
            """Função autonoma que colocará uma escala em vigor"""

            def procurando_escala():
                for esc in os.listdir('../SargenteanteAlpha/Escalas'):
                    if DATA_HOJE == esc.replace('escala', ''):
                        # achamos a escala do dia
                        messagebox.showinfo(message='Executando escala do dia')
                        return esc  # não necessario continuar
                    """
                    DEPOIS DA REESTRUTURAÇÃO, NEM PRECISA MAIS
                    essa forma de visualização não é boa, devemos fazer de outra maneira
                    if f'{DIA_HOJE}' in esc:
                    """
                return None

            escala_hoje = procurando_escala()

            if escala_hoje is not None:
                """Não precisa se preocupar se a função de ver a turma vai pegar a atualização, pois essa função será autonoma,
                        enquanto a de ver a turma será exectuda logo dps."""

                # Devemos colocar essas pessoas de serviços:
                #   Adicionando numa unidade a quantidade V
                #   Informar ao usuário V bastou messagebox
                #   Colocar na lista de descanso
                #   e finalmente apagar a escala do dia

                guarnicao = []
                # so temos o número

                para_servicos_gerais = []

                # numero - tipo

                def pegando_efetivando():

                    # Vamos pegar a guarnicao
                    def pegando():
                        # Apenas precisamos guardar a informação do nome da pessoa
                        with open(rf'Escalas/{escala_hoje}') as esc:
                            for linha_ in esc:
                                linha = linha_.split("-")

                                # Note que estamos pegando o numero
                                guarnicao.append(linha[0])

                                linha[0] = linha[0].replace('\n', '')
                                linha[1] = linha[1].replace('\n', '')

                                para_servicos_gerais.append(linha)

                    pegando()

                    # Vamos aumentar numa unidade a quantidade de serviço
                    def efetivando_base():
                        # Vamos procurar esses números nas listas
                        # Não necessariamente teremos pessoas da mesma turma, então pode ser sanha

                        for arq in os.listdir('../SargenteanteAlpha/Arquivos'):
                            houve_alteracao = False
                            if '.txt' in arq:
                                # Achamos uma base, vamos procurar nela.
                                # Obtendo as linhas da base
                                p = open(f'Arquivos/{arq}', 'r')
                                linhas = p.readlines()
                                p.close()

                                # Para varrermos a lista do Sanha
                                index = 0
                                for linha in linhas:
                                    # Vamos tratar cada linha

                                    if '\n' in linha:
                                        linha = linha.replace("\n", '')
                                        linhas[index] = linhas[index].replace('\n', '')

                                    # Vamos fatorar as informações
                                    linha = linha.split("-")

                                    if linha[0] in guarnicao:
                                        # Este guerreiro está na guarnicao, devemos alterar.
                                        # Lembre que já está sem o \n
                                        linha[4] = str(int(linha[4]) + 1)
                                        # Vamos modificar a linha completa
                                        linhas[index] = '-'.join(linha)

                                        # Como houve apenas uma alteração, vale o Sanha
                                        houve_alteracao = True

                                    index += 1

                                # Supondo que houve alteração em alguma
                                if houve_alteracao:
                                    # Vamos apagar
                                    os.remove(f'Arquivos/{arq}')

                                    # Vamos recriar
                                    primeira = True
                                    with open(f"Arquivos/{arq}", 'x') as base:
                                        for peca in linhas:
                                            if primeira:
                                                base.write(peca)
                                                primeira = False
                                            else:
                                                base.write(f'\n{peca}')

                    efetivando_base()

                pegando_efetivando()

                def descansando():
                    # Vamos abrir e salvar as informações
                    linhas = []
                    with open('../SargenteanteAlpha/SubArquivos/Descansos') as desc:
                        for linha in desc:
                            linhas.append(linha)

                    # Agora que ja salvamos as que existem, precisamos salvar os novos
                    # É garantia que a pessoa em guarnicao não estará no descanso
                    final = seguinte(seguinte(seguinte(obtendo_hoje())))
                    final = final.replace(final[-1], '')
                    for id_pessoa in guarnicao:
                        if len(linhas) == 0:
                            linhas.append(f'{id_pessoa}-{final}')
                        else:
                            linhas.append(f'\n{id_pessoa}-{final}')

                    # agora vamos reescrever o arquivo
                    os.remove('../SargenteanteAlpha/SubArquivos/Descansos')

                    with open('../SargenteanteAlpha/SubArquivos/Descansos', 'x') as de:
                        for precisa_descanso in linhas:
                            de.write(precisa_descanso)

                    # ATENÇÃO, NA ESCALA TEMOS NOME DE GUERRA E NOS DESCANSO TEMOS O NÚMERO DO ALUNO não pode ser diferente...

                descansando()

                def apagando():
                    limpeza(escala_hoje, 0)

                apagando()

                preenchendo_servicos_gerais(para_servicos_gerais)

            else:
                messagebox.showwarning(message='Não há escalas para hoje')

        aplicando_escala()

    def mexendo_baixados():
        """Aqui, vamos centralizar tudo que pudermos sobre o sistema de baixamento.
        Primeiro, devemos ter uma interface basica.
            Vamos mostrar uma lista das pessoas baixadas, mostrando o nome de guerra, o dia que entrou e que sairá.
            Devemos ter um historico mostrando todas as pessoas.
            O motivo de baixamento não é nada que importa tanto, logo devemos ficar atento para vermos o que iremos fazer

            Deveremos ter um arquivo de baixados para mostrar as pessoas que estao baixados no momento e
            um historico de baixados"""

        # Não precisaremos de uma exclusão, vamos incluir todas as pessoas aqui. Vamos adicionar uma nova forma de dado:
        # o número do aluno

        baixados = {'Nome de Guerra': [], 'Entrada': [], 'Saida': [], 'Motivo': []}

        def visualizando_listageral():
            """Aqui, vamos guardar todas as informações sobre os alunos para mostrar.
            Teremos várias planilhas em excell para serem mostradas.
            Tudo vai se basear na lista excell que esta sendo mostrada."""

            # Vamos criar a moldura
            def moldura():
                molde = Frame(menu_baixados, relief='solid', bg='#dde', borderwidth=2)
                molde.place(x=5, y=10, width=230, height=350)
                return molde

            mol = moldura()

            # Deve ser algo geral e autonomo
            def criando_treeview():

                # Vamos zerá-la para não ficar acumulando
                def zerando():
                    baixados['Nome de Guerra'].clear()
                    baixados['Entrada'].clear()
                    baixados['Saida'].clear()
                    baixados['Motivo'].clear()

                zerando()

                # Vamos extrair as informações
                def extraindo_baixados():
                    p_ = 0
                    with open(r'../SargenteanteAlpha/SubArquivos/baixados') as lista:
                        for linha in lista:
                            linha1 = linha.split('-')
                            baixados['Nome de Guerra'].append(linha1[0])
                            baixados['Entrada'].append(linha1[1])
                            baixados['Saida'].append(linha1[2])
                            baixados['Motivo'].append(linha1[3])

                            p_ += 1

                    return p_

                p = extraindo_baixados()

                def colocando_tree():
                    tv = ttk.Treeview(menu_baixados, columns=['Nome de Guerra', 'Entrada', 'Saida'],
                                      show='headings')

                    tv.column('Nome de Guerra', minwidth=0, width=70)
                    tv.heading('Nome de Guerra', text='Nome de Guerra')
                    tv.column('Entrada', minwidth=0, width=50)
                    tv.heading('Entrada', text='Entrada')
                    tv.column('Saida', minwidth=0, width=30)
                    tv.heading('Saida', text='Saida')
                    tv.place(x=18, y=70, height=280, width=200)

                    for c in range(0, p):
                        # print(baixados['Nome Completo'][c])
                        tv.insert("", "end", values=[baixados['Nome de Guerra'][c],
                                                     baixados['Entrada'][c],
                                                     baixados['Saida'][c]])

                colocando_tree()

            criando_treeview()

            def filtro():

                # Vamos precisar criar um combox mostrar as possiveis capacidade de filtro
                def mostrando_filtros():
                    possiveis = ['1°', '2°', '3°', '4°', '5°']

                    filtros_possiveis = ttk.Combobox(mol, values=possiveis)
                    filtros_possiveis['state'] = 'readonly'
                    filtros_possiveis.set(possiveis[0])
                    filtros_possiveis.place(x=10, y=20, width=50)

                    return filtros_possiveis

                filtro_escolhido = mostrando_filtros()

                def aplicando_filtro():
                    filtro1 = filtro_escolhido.get()

                    filtro1 = int(filtro1.replace('°', ''))

                    # Vamos preencer com os nomes a serem vasculhados
                    nomes_guerra = []

                    def vasculhando():
                        # Algoritmo para obter o ano da turma corresponde ao ano
                        ano_a_ser_vasculhado = (ANO_HOJE - 2000) + 5 - filtro1

                        # Vamos precisar abrir o Sanha e pegar os nomes a serem buscados:
                        for arq in os.listdir('../SargenteanteAlpha/Arquivos'):
                            if f'{ano_a_ser_vasculhado}' in arq and '.txt':

                                with open(rf'Arquivos/{arq}') as completo:
                                    for linha_ in completo:
                                        linha = linha_.split("-")
                                        if linha[0] != '\n':
                                            nomes_guerra.append(linha[2])

                    vasculhando()

                    # Vamos ter que fazer o filtro de uma forma inteligente
                    def colocando_tree():
                        tv = ttk.Treeview(menu_baixados, columns=['Nome de Guerra', 'Entrada', 'Saida'],
                                          show='headings')

                        tv.column('Nome de Guerra', minwidth=0, width=70)
                        tv.heading('Nome de Guerra', text='Nome de Guerra')
                        tv.column('Entrada', minwidth=0, width=50)
                        tv.heading('Entrada', text='Entrada')
                        tv.column('Saida', minwidth=0, width=30)
                        tv.heading('Saida', text='Saida')
                        tv.place(x=18, y=70, height=280, width=200)

                        # Essa busca tem que ser feita, pois nós queremos mostrar os nomes do respectivo ano selecionado

                        try:

                            # Função para verificar se o nome está presente na lista
                            def buscador(indice):
                                if baixados['Nome de Guerra'][indice] in nomes_guerra:
                                    return True
                                else:
                                    return False

                            k = 0
                            while True:

                                if buscador(k):
                                    tv.insert("", "end", values=[baixados['Nome de Guerra'][k],
                                                                 baixados['Entrada'][k],
                                                                 baixados['Saida'][k]])

                                k += 1

                        except IndexError:
                            pass

                    colocando_tree()

                Button(mol, text='Obter', command=aplicando_filtro).place(x=80, y=20, height=20)

                informador(mol, 130, 17, 'Mostrará a lista de baixados do respectivo ano')

            filtro()

        visualizando_listageral()

        def motivo_baixado():
            moti = LabelFrame(menu_baixados, text='Motivo de Baixamento', bg="#dde", relief='solid')
            moti.place(x=250, y=10, width=330, height=100)

            def pegador_motivo():
                """Vai entrar na lista de baixados e pegar o motivo"""

                # Vamos ter que criar uma janela para usá-las
                def janela_nova():

                    sanha1 = criando_janela('Motivador', 300, 200)

                    p = Frame(sanha1, relief='solid', bg='#dde', borderwidth=2)
                    p.place(x=5, y=5, width=290, height=190)

                    Label(p, text='NOME: ', bg='#dde').place(x=20, y=10)
                    nome1 = Entry(p)
                    nome1.place(x=70, y=10, width=100)

                    Label(p, text='Press Enter', bg='#dde').place(x=200, y=10)

                    return [sanha1, p, nome1]

                sanha, frame, nome = janela_nova()

                def motivador(event):

                    # Vamos pegar o nome carteado
                    NOME = nome.get().upper().strip()

                    # Vamos verificar se existe algum nome na lista, sem nem precisar digitar o nome completo

                    def colocando_janela():
                        c = 0
                        try:
                            for aluno in baixados['Nome de Guerra']:
                                if verificador(NOME, aluno):
                                    # Vamos obter daqui o indice
                                    break

                                c += 1

                            delta = 20
                            """Um bug que estavamos tendo é que o nome do aluno apenas subtituido, por isso uma parte do 
                            anterior ainda era visivel e ficava no sanha"""
                            Label(frame, text=f'ALUNO: {baixados["Nome de Guerra"][c]}', bg='#dde').place(x=60,
                                                                                                          y=40 + delta,
                                                                                                          width=150)
                            Label(frame,
                                  text=f'ENTRADA: {baixados["Entrada"][c]}' + ' ' * 35 + f'SAIDA: {baixados["Saida"][c]}',
                                  bg='#dde').place(x=15, y=70 + delta)

                            texto = baixados["Motivo"][c]

                            Label(frame, text=f'MOTIVO: {texto}', bg='#dde').place(x=15, y=100 + delta)

                        except IndexError:
                            messagebox.showerror(title='Sem Lista Selecionada',
                                                 message="""Por favor, visualize a lista antes ou escreva o nome de forma correta""")

                    colocando_janela()

                nome.bind('<Return>', motivador)

                sanha.mainloop()

            Button(moti, text='Obter Motivo de Baixamento', command=pegador_motivo).place(x=80, y=20)

            informador(moti, 250, 20,
                       'A partir de um dado do aluno, mostrará o motivo, data inicial e final de seu respectivo baixamento.')

        motivo_baixado()

        def baixando():
            """Aqui vamos reunir os nossos fundamentos para o ato de baixar alguem"""

            # Só a moldura
            sanha = LabelFrame(menu_baixados, text='Baixando', bg='#dde', borderwidth=2, relief='solid')
            sanha.place(x=250, y=120, width=330, height=100)

            def baixando_guerreiro():
                comprimento = 300
                altura = 350

                baix = criando_janela('', comprimento, altura)

                ba = Frame(baix, borderwidth=2, relief='solid', bg='#dde')
                ba.place(x=10, y=10, width=comprimento - 20, height=altura - 20)

                # Pegando o número da pessoa baixada
                Label(ba, text="NUMERO ALUNO: ", bg='#dde').place(x=10, y=10)
                num = Entry(ba)
                num.place(x=115, y=10, width=60)

                # Pegando a data que entrou de baixado
                def pegando_entrada():
                    Label(ba, text='ENTRADA(DIA-MÊS): ', bg="#dde").place(x=10, y=40)
                    dia_carteado = StringVar(ba)
                    dia_carteado.set(f"{DATA_HOJE.replace(DATA_HOJE[-1], '')}")
                    dia_entrada1 = Spinbox(ba, from_=1, to=31, textvariable=dia_carteado)
                    dia_entrada1.place(x=130, y=40, width=30)
                    mes_carteado = StringVar(ba)
                    mes_carteado.set(f"{codigo_mes(DATA_HOJE[-1], False)}")
                    mes_entrada1 = Spinbox(ba, from_=1, to=12, textvariable=mes_carteado)
                    mes_entrada1.place(x=170, y=40, width=30)
                    return [dia_entrada1, mes_entrada1]

                dia_entrada, mes_entrada = pegando_entrada()

                # Pegando a data que teoricamente sairá de baixado
                def pegando_saida():
                    Label(ba, text='SAIDA(DIA-MÊS): ', bg="#dde").place(x=10, y=70)
                    dia_saida1 = Spinbox(ba, from_=1, to=31)
                    dia_saida1.place(x=110, y=70, width=30)
                    mes_saida1 = Spinbox(ba, from_=1, to=12)
                    mes_saida1.place(x=150, y=70, width=30)
                    return [dia_saida1, mes_saida1]

                dia_saida, mes_saida = pegando_saida()

                # Pegando o motivo do baixamento
                Label(ba, text='MOTIVO: ', bg="#dde").place(x=10, y=120)
                Motivo = Text(ba)
                Motivo.place(x=80, y=120, width=150, height=120)

                def incapacitando():

                    try:
                        numero2 = num.get()
                        # Vamos obter as informações no sanha
                        nome = pegando_info(numero2, [0, 2])
                        # print(nome)

                        if nome is False:
                            # Caso dê um erro
                            return messagebox.showwarning(message='Aluno não encontrado.')

                        def verificar_se_ja_presente():
                            p = open('../SargenteanteAlpha/SubArquivos/baixados')
                            linhas = p.readlines()
                            p.close()

                            for linha in linhas:
                                linha = linha.split('-')
                                if linha[0] == nome:
                                    return True

                        if verificar_se_ja_presente():
                            return messagebox.showwarning(message='Aluno já está baixado nesse momento.')

                        entrada = f"{dia_entrada.get()}/{mes_entrada.get()}"
                        # print(entrada)

                        saida = f"{dia_saida.get()}/{mes_saida.get()}"
                        # print(saida)

                        motivo = Motivo.get(1.0, "end-1c").upper()

                        # print(motivo)

                        # Vamos escrever as informações tanto no baixados atualmente como no histórico

                        def escrevendobaixados():

                            # Verificando se está vazio
                            arq = open(r'../SargenteanteAlpha/SubArquivos/baixados', 'r')
                            q = len(arq.readlines())
                            arq.close()

                            arq = open(r'../SargenteanteAlpha/SubArquivos/baixados', 'a')

                            if q == 0:
                                # Se não há ninguem vamos precisar inserir do jeito certo
                                arq.write(f'{nome}-{entrada}-{saida}-{motivo}')
                            else:
                                # sanha safo
                                arq.write(f'\n{nome}-{entrada}-{saida}-{motivo}')

                            arq.close()

                        escrevendobaixados()

                        def escrevendohistorico():
                            # Verificando se está vazio
                            arq = open(r'../SargenteanteAlpha/SubArquivos/historicobaixados', 'r')
                            q = len(arq.readlines())
                            arq.close()

                            arq = open(r'../SargenteanteAlpha/SubArquivos/historicobaixados', 'a')

                            if q == 0:
                                # Se não há ninguem vamos precisar inserir do jeito certo
                                arq.write(f'{nome}-{entrada}-{saida}-{motivo}')

                            else:
                                # sanha safo
                                arq.write(f'\n{nome}-{entrada}-{saida}-{motivo}')

                            arq.close()

                        escrevendohistorico()

                        def verificando_escalas():
                            presente = []
                            for arquivo in os.listdir('../SargenteanteAlpha/Escalas'):
                                with open(f'Escalas/{arquivo}') as verif:
                                    for linha_ in verif:
                                        if num.get() in linha_:
                                            presente.append(arquivo)
                                            # Se achamos ele na escala, não há necessidade de continuar a procurar ele
                                            # na mesma escala ne
                                            break

                            if len(presente) != 0:
                                for escal in presente:
                                    messagebox.showwarning(title='SANHA',
                                                           message=f'Baixado foi encontrado na {escal}')

                                # Deixo aqui os meus pesâmes, pois foi tentado criar uma forma de automatizar as substituições
                                # devido ao baixado, aqui jaz as Sanhas:
                                # Os plantões são bizarros, a forma de subir depende do ano e dos tipos de serviço
                                # temos que automaticamente destruir a última escala e recriá-la
                                # Mano, simplesmente não dá. Desiste disso

                        verificando_escalas()
                    except (TypeError, IndexError):
                        messagebox.showerror(title='ERROR',
                                             message="Informações Inseridas não são válidas, tente revê-las")

                    visualizando_listageral()
                    baix.destroy()

                Button(ba, text='Baixá-lo', command=incapacitando).place(x=100, y=270)

                baix.mainloop()

            Button(sanha, text='Baixando Guerreiro', command=baixando_guerreiro).place(x=100, y=20)

            informador(sanha, 220, 20, 'Obterá todas as informações necessárias para baixar alguém.')

        baixando()

        def gerador_planilha():
            """ Aqui vamos construir a função que criará a planilha que desejamos."""

            # Só a moldura
            sanha = LabelFrame(menu_baixados, text='Gerando Planilha', bg='#dde', borderwidth=2, relief='solid')
            sanha.place(x=250, y=230, width=330, height=100)

            def planilhando():

                # Primeiro, devemos obter de qual arquivo se deve importar
                jan = criando_janela('Qual Arquivo Gerar', 270, 100)

                Label(jan, text='Escolha qual arquivo deseja gerar a planilha: ', bg='#dde').place(x=10, y=15)

                def planilhado(opcao):
                    nome = 'b' + 'aixados.xlsx'
                    if opcao == 1:
                        # vamos gerar o momento atual
                        pass
                    else:
                        # vamos gerar o historico
                        nome = 'h' + 'istoricobaixados.xlsx'

                    # Vamos pegar as informações
                    def extraindo():
                        # vamos entrar e pegar as linhas
                        linhas = []
                        with open(f'SubArquivos/{nome.replace(".xlsx", "")}', 'r') as arq:
                            for linha in arq:
                                if '\n' in linha:
                                    linha = linha.replace('\n', '')

                                linhas.append(linha.split('-'))

                        return linhas

                    dados = extraindo()
                    # temos aqui [[nome guerra, entrada, saida, motivo]

                    # Agora, vamos criar o Sanha

                    try:
                        import openpyxl as xl
                        from openpyxl.styles import Font

                        # Afinal, vamos criar a planilha
                        pla = xl.Workbook()

                        pagina = pla.active

                        # É interessante pensar que se o histórico for muito grande, o Sanha irá consumir a velocidade
                        dados_completos = [[pegando_info(pessoa[0], [2, 0]),
                                            pessoa[0],
                                            pessoa[1],
                                            pessoa[2],
                                            pessoa[3]
                                            ] for pessoa in dados]
                        titulos = ['Número', 'Nome de Guerra', 'Entrada', 'Saida', 'Motivo']
                        colunas = ['A', 'B', 'C', 'D', 'E']

                        posicao = 0
                        while True:
                            if posicao == len(titulos):
                                break

                            # Criando o titulo
                            pagina[colunas[posicao] + '1'] = titulos[posicao]
                            pagina[colunas[posicao] + '1'].font = Font(bold=True)

                            # Colocando os valores em cada Célula
                            i = 2
                            for info in dados_completos:
                                pagina.column_dimensions[colunas[posicao]].width = len(info[posicao]) + 10
                                pagina[colunas[posicao] + f'{i}'] = info[posicao]
                                i += 1

                            posicao += 1

                        pla.save(nome)

                    except:
                        return messagebox.showerror(message='Erro ao tentar criar planilha de baixados')

                    jan.destroy()

                Button(jan, text='Momento Atual', command=lambda: planilhado(1)).place(x=10, y=60)
                Button(jan, text='Historico', command=lambda: planilhado(2)).place(x=150, y=60)

                jan.mainloop()

            Button(sanha, text='Planilhando', command=planilhando).place(x=120, y=20)

            informador(sanha, 200, 20,
                       'Vai criar uma planilha .xlsx de um dos arquivos de baixados, seja o do momento atual ou do histórico.')

        gerador_planilha()

        def verificador_baixados():
            # Devemos criar a função capaz de retirar os baixados que chegarão ao dia final.
            # Note que quem está baixado até o dia x, ainda está baixado no dia x, saindo apenas no dia x + 1

            try:
                # Vamos extrair as informações
                p = open('../SargenteanteAlpha/SubArquivos/baixados')
                linhas = p.readlines()
                p.close()

                # De posse desse Sanha, devemos ver que já está fora
                deve_sair = []
                for linha in linhas:
                    linha1 = linha.split('-')
                    # Podemos fazer do jeito pau ou do jeito carteado.
                    data_final = linha1[2].split('/')  # [dia, mes numerico]
                    data_final = f'{data_final[0]}{codigo_mes(int(data_final[1]), True)}'
                    # Dai, a verificação
                    if seguinte(data_final) == DATA_HOJE:
                        messagebox.showinfo(title='Safou', message=f'Aluno {linha1[0]} terminou seu período baixado.')
                        # Vamos salvar a linha que deve sair
                        deve_sair.append(linha)

                for deve_apagada in deve_sair:
                    linhas.remove(deve_apagada)

                os.remove('../SargenteanteAlpha/SubArquivos/baixados')

                with open("../SargenteanteAlpha/SubArquivos/baixados", 'x') as ba:
                    for linha in linhas:
                        ba.write(linha)
            except:
                messagebox.showerror(message='Erro ao verificar se alguém chegou ao final do período baixado.')

        verificador_baixados()

    def mexendo_punidos():
        """Aqui, poderiamos ter algo relacionado aos punidos?
        Vamos organizar como estará organizada as informações
        punidos['ID'] - vai guardar a id de cada aluno punidos
        punidos['punições'] vai ter uma lista inteira composta por várias listas binarias:
                        =  [ ['DRS1', 'SANHUDO DEMAIS'],
                             [...],
                           ]
        Para referenciarmos juntos, vamos guardar assim,
        o indice k dos punidos['ID'] vai guardar tanto o id do aluno quanto as respectivas punições deles, estando
        essas em uma lista

        """

        # Vamos criar a moldura
        def moldura():
            bele1 = Frame(menu_punidos, borderwidth=2, relief="solid", bg="#dde")
            bele1.place(x=5, y=5, width=230, height=365)
            return bele1

        bele = moldura()

        # Organizar as informações assim
        punidos = {'ID': [], 'punicoes': [], 'turma': 0}

        def mostrando_ja_punidos():
            """"""
            """Nessa função, vamos tentar criar um treeview mostrando as pessoas que já foram punidas em algum momento.
            Vamos priorizar sempre mostrar o 1° ano e vamos disponibilizar um Combobox para obter a intenção do usuário."""

            # Função para retornar a turma desejada
            def turmandos(ano):
                return (ANO_HOJE - 2000) + 5 - ano

            # Vamos extrair as informações a partir do ano desejado
            def extraindo_dados(turma_, primeiravez):
                try:

                    # Zerando
                    punidos['ID'].clear()
                    punidos['punicoes'].clear()
                    punidos['turma'] = turma_

                    # vamos abrir o arquivo desejado
                    k = 0  # variavel de controle
                    with open(rf'Punicoes/punidos{turma_}') as puni:
                        for linha_ in puni:

                            if linha_ != '\n':
                                # Vamos retirar esse Sanha
                                linha_ = linha_.replace('\n', '')

                                # Vericando se chegamos a um valor numero, pois indicará que chegamos a um punido
                                if linha_.isnumeric():
                                    punidos['ID'].append(linha_)
                                    punidos['punicoes'].append([])
                                    # O else garantirá que estamos presos a um punido
                                else:

                                    # Verificando se chegamos ao final
                                    if linha_ != '}':
                                        linha = linha_.split(">")

                                        # Guardando as informações em lista
                                        punidos['punicoes'][k].append([linha[0], linha[1]])

                                    else:
                                        # Assim, podemos ter uma lista de lista de punições para cada pessoa
                                        k += 1

                    if k == 0 and (ANO_HOJE - 2000 + 5 - turma_) != 1:
                        messagebox.showwarning(title='AVISO', message='Não há punidos nesse ano')

                except FileNotFoundError:
                    if primeiravez:
                        pass
                    else:
                        messagebox.showwarning(title='AVISO', message='Não há arquivos de punidos para esse ano')

            extraindo_dados(turmandos(1), True)

            # Após extrairmos os dados, devemos preparar e mostrar o treeview
            def mostrando_treeview(turma):

                def tree_view(lista_punidos):
                    # Criando o TreeView
                    tv = ttk.Treeview(bele, columns=['ID', 'Nome de Guerra', 'Quant'],
                                      show='headings')

                    # Colocando as colunas do treeview
                    tv.column('ID', minwidth=0, width=40)
                    tv.heading('ID', text='ID')

                    tv.column('Nome de Guerra', minwidth=0, width=90)
                    tv.heading('Nome de Guerra', text='Nome de Guerra')

                    tv.column('Quant', minwidth=0, width=30, anchor='center')
                    tv.heading('Quant', text='Quant')

                    tv.place(x=5, y=80, height=270, width=215)
                    Label(bele, text='Mostrando Lista de Alunos Já Punidos', bg='#dde').place(x=5, y=60)

                    # Aplicando as informações do TreeView
                    for ID, pessoa, quantidade in lista_punidos:
                        tv.insert("", "end", values=[ID,
                                                     pessoa,
                                                     quantidade])

                # lista que guarda listas compostas por id, nome completo e quantidade.
                lista = []

                # Vamos preeche-la com as informações
                def preenchendo():
                    # Vamos preeche-la com as informações
                    try:
                        k = 0  # variavel de controle

                        while True:
                            numero0 = punidos['ID'][k]
                            quant = len(punidos['punicoes'][k])
                            nome = pegando_info(str(numero0), [0, 2], rf'Arquivos/CFG-{turma}.txt')

                            if nome is None:
                                # Para evitar que alguem varado entre na lista
                                messagebox.showwarning(title='AVISO',
                                                       message=f'Há uma pessoa com um número errado{numero0}')

                            else:

                                # Vamos verificar se é FO POSITIVO
                                for nivel, motivo in punidos['punicoes'][k]:
                                    if nivel == 'Positivo':
                                        # Vamos só ignorar
                                        pass
                                    else:
                                        lista.append([numero0, nome, quant])

                            # Assim vamos garantir que não persista no erro
                            k += 1

                    except IndexError:
                        pass

                preenchendo()

                # Agora precisamos ordenar baseando-se na quantidade de punições
                lista.sort(reverse=True, key=lambda sublista: sublista[2])

                tree_view(lista)

            mostrando_treeview(turmandos(1))

            # Vamos construir um combobox com a opção de escolher os anos e filtrar
            def anando():

                # Criando Combobox
                def combobox_():
                    possiveis = ['1°', '2°', '3°', '4°', '5°']

                    filtros_possiveis = ttk.Combobox(bele, values=possiveis)
                    filtros_possiveis['state'] = 'readonly'
                    filtros_possiveis.set(possiveis[0])
                    filtros_possiveis.place(x=10, y=30, width=50)

                    return filtros_possiveis

                ano_desejado = combobox_()

                # Devemos criar o botão e a respectiva função
                def filtrando():
                    ano = int(ano_desejado.get().replace('°', ''))

                    extraindo_dados(turmandos(ano), False)
                    mostrando_treeview(turmandos(ano))

                Button(bele, text='Obter', command=filtrando).place(x=80, y=30, height=20)

                informador(bele, 130, 27,
                           'Mostrará a lista de todos os que JÁ FORAM punidos do respectivo ano')

            anando()

        mostrando_ja_punidos()

        def mostrando_punicoes():
            # Vamos procurar no arquivo a partir do nome e mostrar todas as punições da pessoa com as suas respectivas
            # punições

            def criando_espaco():
                most = LabelFrame(menu_punidos, text='Mostrando Punições', relief='solid', borderwidth=2, bg='#dde')
                most.place(x=250, y=5, width=335, height=100)

                informador(most, 85, 20, 'Clicando duas vezes em uma punição, você consegue apagar ela.', ' - ')

                Button(most, text='Visualizar', command=obter_punicoes).place(x=120, y=20, width=100)

                informador(most, 230, 20,
                           'A partir do número, vai mostrar todas as punições registradas e suas respectivas descrições')

            def obter_punicoes():
                # Vamos ter que criar uma janela e disponibilizá-la para pegar o Sanha

                puni = criando_janela("Mostrando Punições", 400, 250)

                # A informação do usuário
                def informa():
                    Label(puni, text='Insira o número:', bg='#dde').place(x=40, y=10)
                    inf = Entry(puni)
                    inf.place(x=145, y=10, width=80)
                    return inf

                infor = informa()

                # Botão para pegar a informação
                def buscar(event):
                    num = infor.get()

                    # Vamos guardar as informações
                    punicoes = []

                    # Agora, devemos procurar o arquivo e mostrar as punições
                    def procurando_punido():
                        try:
                            k = 0
                            for numero9 in punidos['ID']:
                                # Tentando achar

                                """Um erro que estavamos tendo é que num é string, logo oq devemos fazer é int() e não f'{num}'"""
                                if num == numero9:

                                    # Agora que achamos o aluno, vamos pegar as infor

                                    for niv, pun in punidos['punicoes'][k]:
                                        # Vamos salvar os dados
                                        punicoes.append([niv, pun])

                                    # Não precisamos vasculhar mais
                                    return 0

                                k += 1

                            messagebox.showerror(title='ERROR', message='O aluno não está presente nas punições')
                        except:
                            messagebox.showerror(title='ERROR', message='O aluno não está presente nas punições')

                    procurando_punido()

                    def treeview_final(lista, mestre):

                        try:
                            # Criando o TreeView
                            tv = ttk.Treeview(mestre, columns=['Nível', 'Descrição'],
                                              show='headings')

                            # Colocando as colunas do treeview
                            tv.column('Nível', minwidth=0, width=5, anchor='center')
                            tv.heading('Nível', text='Nível')

                            tv.column('Descrição', minwidth=0, width=300, anchor='center')
                            tv.heading('Descrição', text='Descrição')

                            tv.place(x=10, y=40, width=380, height=200)

                            # Aplicando as informações do TreeView
                            for nivel, descri in lista:
                                # print(informacoes['Nome Completo'][c])
                                tv.insert("", "end", values=[nivel, descri])

                            def apagar_punicao(evet):
                                if messagebox.askyesno(message='Tem certeza que deseja apagar essa punição?'):
                                    try:
                                        item = tv.focus()

                                        punicao = tv.item(item)['values']

                                        turma = pegando_info(num, [0, -1])

                                        # Vamos primeiro apagar no treeview
                                        tv.delete(item)

                                        # Vamos apagar no arquivo
                                        p = open(f'Punicoes/punidos{turma}', 'r')
                                        # Note que vai ter o \n com certeza
                                        linhas = [lin.replace('\n', '') for lin in p.readlines()]
                                        p.close()

                                        # Vamos agora retirar o sanha do Sanha
                                        # Em teoria, poderiamos apenas colocar .remove('>'.join(punicao))
                                        # Mas e se o cara tiver só uma punição? Devemos procurar a lógica
                                        def apagando_decentemente(numero_safo, pun):
                                            def sabendo_o_que_restara():
                                                # Linhas a ser apagadas
                                                a_ser_apagada = []

                                                index = 0
                                                for caixa in linhas:
                                                    # Vamos primeiro achar o guerreiro pelo número

                                                    if numero_safo == caixa:

                                                        # Achamos o sanhudo
                                                        j = 1
                                                        quantidade_alem_da_especifica = -1
                                                        while True:

                                                            a_ser_apagada.append(linhas[index + j - 1])

                                                            # Verificando se chegamos ao final do sanhudo.
                                                            if linhas[index + j - 1] == '}':
                                                                # Chegamos ao final
                                                                # Se chegamos ao final
                                                                return a_ser_apagada, quantidade_alem_da_especifica

                                                            # Em teoria, vamos comecar verificando quantas punicoes esse sanhudo tem
                                                            # Devemos contar quantas punicoes ele tem além da sanhuda
                                                            if linhas[index + j] == pun:
                                                                pass
                                                            else:
                                                                quantidade_alem_da_especifica += 1

                                                            j += 1

                                                    index += 1

                                            linhas_a_serem_apagadas, quant_alem_da_especi = sabendo_o_que_restara()

                                            def apagando_de_fato():
                                                # Vamos verificar se o sanhudo possui apenas uma punição, caso sim:
                                                if quant_alem_da_especi == 0:
                                                    # Vamos apagar ele inteiro do sanha
                                                    for linha_a_ser_apagada in linhas_a_serem_apagadas:
                                                        linhas.remove(linha_a_ser_apagada)

                                                else:
                                                    # Como há mais de uma punição, vamos sanhar
                                                    linhas.remove(pun)

                                            apagando_de_fato()

                                            def reescrevendo(linhas1):
                                                linhas1 = [linha for linha in linhas1 if linha != '']
                                                # Vamos reescrever o Sanha

                                                quant = len(linhas1) - 1

                                                os.remove(f'Punicoes/punidos{turma}')

                                                with open(f'Punicoes/punidos{turma}', 'x') as lista_punidos:
                                                    index = 0
                                                    for linha in linhas1:
                                                        # Em teoria, vai ter um \n EM TUDO, menos no Sanha final

                                                        if index == quant:
                                                            lista_punidos.write(linha)
                                                        else:
                                                            lista_punidos.write(f'{linha}\n')

                                                        index += 1

                                            reescrevendo(linhas)

                                            puni.destroy()
                                            mostrando_ja_punidos()

                                        apagando_decentemente(num, '>'.join(punicao))

                                        if len(linhas) == 0:
                                            mestre.destroy()
                                    except:
                                        pass

                            # Vamos construir o Sanha
                            tv.bind("<Double-1>", apagar_punicao)

                        except:
                            pass

                    treeview_final(punicoes, puni)

                infor.bind('<Return>', buscar)
                Label(puni, text='Press Enter', bg='#dde').place(x=240, y=10)

                def executando_janela(puni1):
                    puni1.mainloop()

                executando_janela(puni)

            criando_espaco()

        mostrando_punicoes()

        def pegando_punicao():
            # Vamos ter que construir a moldura e o botão preparado

            def criando_espaco():
                most = LabelFrame(menu_punidos, text='Aplicando Punição', relief='solid', borderwidth=2, bg='#dde')
                most.place(x=250, y=120, width=335, height=100)

                Button(most, text='Aplicar FO em Aluno', command=punir).place(x=100, y=20, width=150)

                informador(most, 260, 20, 'Obterá todas as informações necessárias para registrar uma punição.')

            def punir():

                puni = criando_janela("Pegando Punições", 400, 250)

                def colocando_opcoes():
                    Label(puni, text='Insira o número:', bg='#dde').place(x=10, y=10)
                    inf_ = Entry(puni)
                    inf_.place(x=110, y=10, width=80)

                    Label(puni, text='Insira o ano:', bg='#dde').place(x=220, y=10)
                    ano__ = Spinbox(puni, from_=1, to=5)
                    ano__.place(x=300, y=10, width=30)

                    Label(puni, text='Insira a descrição(Tente inserir data):', bg='#dde').place(x=10, y=50)
                    descri__ = Text(puni)
                    descri__.place(x=10, y=70, width=200, height=150)

                    Label(puni, text='Selecione o nível de punição:', bg='#dde').place(x=230, y=50)
                    possiveis_punicoes = ['DRS1', 'DRS2', 'II', 'IA', 'FATD', 'Positivo']
                    pu = ttk.Combobox(puni, values=possiveis_punicoes)
                    pu['state'] = 'readonly'
                    pu.place(x=230, y=80, width=150)
                    pu.set(possiveis_punicoes[0])

                    return [ano__, descri__, inf_, pu]

                ano_, descri, inf, niv = colocando_opcoes()

                def salvando():
                    # Vamos vasculhar os anos de base que existem para achar a pessoa
                    # Motivo.get(1.0, "end-1c").upper()

                    turma_ = (ANO_HOJE - 2000) + 5 - int(ano_.get())

                    descricao = descri.get(1.0, "end-1c").upper()

                    num = inf.get()

                    nivel = niv.get()

                    # Antes de qualquer coisa, devemos verificar se o aluno existe nesse ano
                    def fazendo_verificacao():

                        try:
                            p = pegando_info(num, [0, -1], rf'Arquivos/CFG-{turma_}.txt')

                            if p:
                                # quer dizer que não, não existe
                                return False
                            else:
                                # quer dizer que não existe
                                return True

                        except FileNotFoundError:
                            messagebox.showerror(title='ERROR', message='Esse ano não possui base de arquivo')
                            puni.destroy()

                    nao_existe = fazendo_verificacao()

                    if nao_existe:
                        messagebox.showerror(title='ERROR',
                                             message='O referido número não representa um aluno do referido ano')
                        puni.destroy()
                        return 0

                    torrando(num, nivel, descricao, turma_)

                    puni.destroy()
                    mostrando_ja_punidos()

                Button(puni, text='Aplicar', command=salvando).place(x=250, y=150, height=20, width=100)

                """Não da para colocarmos o imprimr aqui pois nós só queremos que seja extraido apos pegar os dados"""

                puni.mainloop()

            criando_espaco()

        pegando_punicao()

        def carregando_punicoes():
            # A partir de uma planilha, vai pegar todas as punições de cada um dos alunos,
            # Inclusive, poderá receber até um PDF.
            # Vamos pensar em 2 formas de obtenção, PDF ou planilha.
            # Sendo assim, vamos construir um método para obter a planilha primeiro e depois construiremos para receber o PDF

            def para_extrair():
                try:
                    def porplanilha(mestre):
                        mestre.destroy()

                        # Primeiro, devemos verificar de qual planilha falamos

                        def obtendo_planilha():

                            possivel_base = []
                            for arq in os.listdir(os.getcwd()):
                                # print(arq)
                                if '.xlsx' in arq:
                                    possivel_base.append(arq)

                            if len(possivel_base) == 0:
                                return messagebox.showwarning(message='Não nenhuma planilha que se pode obter punidos.')

                            jan1 = criando_janela('Escolha a base', 200, 100)

                            cb = ttk.Combobox(jan1, values=possivel_base)
                            cb['state'] = 'readonly'
                            cb.set(possivel_base[0])
                            cb.place(x=10, y=10)

                            def escolher():
                                extrair_planilha(cb.get())
                                jan1.destroy()
                                mostrando_ja_punidos()

                            Button(jan1, text='Pronto', command=escolher).place(x=10, y=40)

                            jan1.mainloop()

                        obtendo_planilha()

                    def porpdf(mestre):
                        mestre.destroy()

                        def obtendo_pdf():

                            possivel_base = []
                            for arq in os.listdir(os.getcwd()):
                                # print(arq)
                                if '.pdf' in arq:
                                    possivel_base.append(arq)

                            if len(possivel_base) == 0:
                                return messagebox.showwarning(
                                    message='Não há nenhum arquivo pdf para se extrair punidos')

                            jan1 = criando_janela('Escolha a base', 200, 100)

                            cb = ttk.Combobox(jan1, values=possivel_base)
                            cb['state'] = 'readonly'
                            cb.set(possivel_base[0])
                            cb.place(x=10, y=10)

                            def escolher():
                                extrair_PDF(cb.get())
                                jan1.destroy()
                                mostrando_ja_punidos()

                            Button(jan1, text='Pronto', command=escolher).place(x=10, y=40)

                            jan1.mainloop()

                        obtendo_pdf()

                    def obtendo_metodo_de_extracao():
                        jan = criando_janela('Método de Extração', 300, 300)

                        Button(jan, text='Por Planilha', borderwidth=5, command=lambda: porplanilha(jan)).place(x=0,
                                                                                                                y=0,
                                                                                                                width=300,
                                                                                                                height=150)

                        Button(jan, text='Por PDF', borderwidth=5, command=lambda: porpdf(jan)).place(x=0, y=150,
                                                                                                      width=300,
                                                                                                      height=150)

                        jan.mainloop()

                    obtendo_metodo_de_extracao()
                except:
                    return messagebox.showerror(message='Erro na extração de punidos')

            def gerar_punidos():
                # Em teoria, vamos pegar as punicoes de cada um e gerar o Sanha enlatado
                # Primeiro, devemos saber de qual turma estamos falando
                def pegando_ano():
                    return (ANO_HOJE - 2000) + 5 - punidos['turma']

                if messagebox.askyesnocancel(
                        message=f'Você está prestes a gerar uma planilha dos punidos do {pegando_ano()} ano.' +
                                'Você pode alterar o ano apenas filtrando ao lado esquerdo.'):

                    # Já temos tudo que precisamos, vamos apenas iniciar o Sanha de gerar a planilha

                    # Vamos criar ela
                    planilha = xl.Workbook()
                    nome = f'punidos_{punidos["turma"]}.xlsx'

                    # Vamos construir o Sanha
                    pagina = planilha.active

                    # Como em teoria haverá vários punidos, devemos conseguir mostrar todas as punições de cada um deles
                    # E de forma inteligente.
                    cor_punicao = PatternFill(start_color='00FFFF00', end_color='00FFFF00',
                                              fill_type='solid')
                    cor_espaco = PatternFill(start_color='00C0C0C0', end_color='00C0C0C0',
                                             fill_type='solid')

                    index_aluno = 0
                    index_linha = 2
                    for numero_aluno in punidos['ID']:
                        if index_aluno == 0:
                            # Como é o primeiro, vamos iniciar bradando o titulo das colunas
                            pagina['A1'] = 'Número'
                            pagina['A1'].font = Font(bold=True)
                            pagina['B1'] = 'Nome'
                            pagina['B1'].font = Font(bold=True)
                            pagina['C1'] = 'Nível de Punição'
                            pagina['C1'].font = Font(bold=True)
                            pagina.column_dimensions['C'].width = len('Nível de Punição')
                            pagina['D1'] = 'Punição'
                            pagina['D1'].font = Font(bold=True)
                            pagina.column_dimensions['D'].width = 30

                        # Vamos colocar as informações primordiais de cada um
                        pagina['A' + str(index_linha)] = numero_aluno

                        pagina['B' + str(index_linha)] = pegando_info(numero_aluno, [0, 2])

                        # Vamos precisar montar a lógica das punições
                        for nivel, punicao in punidos['punicoes'][index_aluno]:
                            pagina['C' + str(index_linha)] = nivel
                            pagina['C' + str(index_linha)].fill = cor_punicao
                            pagina['D' + str(index_linha)] = punicao
                            pagina['D' + str(index_linha)].fill = cor_punicao
                            index_linha += 1

                        # Como terminamos nesse aluno, vamos deixar uma barra de espaço
                        for letra in ['A', 'B', 'C', 'D']:
                            pagina[letra + str(index_linha)].fill = cor_espaco

                        index_aluno += 1
                        index_linha += 1

                    planilha.save(nome)

            def criando_espaco():
                most = LabelFrame(menu_punidos, text='Ferramentas de Grande Contingente para Punidos', relief='solid',
                                  borderwidth=2, bg='#dde')
                most.place(x=250, y=235, width=335, height=100)

                pos1 = 180
                Button(most, text='Extrair Punidos', command=para_extrair).place(x=pos1, y=10, width=100)

                informador(most, pos1 + 110, 10, 'Disponibilizará formas de extrair grandes quantidades de punidos.')

                pos2 = 60
                Button(most, text='Gerar Punidos', command=gerar_punidos).place(x=pos2, y=40, width=100)

                informador(most, pos2 - 30, 40, 'Gerará uma planilha de punidos.')

            criando_espaco()

        carregando_punicoes()

    def buscador_infinito():
        # Criando a função que fará o processo de buscar

        def buscador():
            messagebox.showinfo(
                message='Com essa função, você é capaz de encontrar qualquer aluno com qualquer tipo de informação.')

            # Primeiro, uma janela deve ser criada e deve ser inserido o tipo de informação que temos e ela mesma

            # Vamos precisar definir a função de mostrar todos os dados aqui, pois dá muito ruim esse ngc
            # de criar um botão para usar uma função
            def finalizando(guerreiro):

                """{'Nome Completo': [], 'Nome de Guerra': [], 'Seg': [], 'Serv': [], 'ID': [], 'Turma': []
                             'Punicoes': [],
                             'Escalas': []}"""

                def pegando_escalas_presentes():
                    # primeiro, a partir do nome do guerreiro, vamos buscá-lo nas escalas
                    for arquivo in os.listdir('../SargenteanteAlpha/Escalas'):
                        # Vamos procurar nas escalas
                        with open(rf'Escalas/{arquivo}') as esca:
                            for linha_ in esca:
                                linha = linha_.split("-")

                                # Quando acharmas, vamos adicionar
                                if guerreiro['ID'] == linha[0]:
                                    guerreiro['Escalas'].append(arquivo)
                                    # Não precisamos varrer mais
                                    break

                pegando_escalas_presentes()

                def pegando_punicoes():
                    # Agora nas punições
                    c = False
                    for arquivo in os.listdir('../SargenteanteAlpha/Punicoes'):
                        with open(rf'Punicoes/{arquivo}') as lista:
                            for linha in lista:
                                # Vamos eliminar esse \n devido as comparações
                                linha = linha.replace('\n', '')
                                if linha.isnumeric():
                                    # confirmando que será um número, poderemos verificar
                                    if linha == guerreiro['ID']:
                                        # Agora que confirmamos, precisamos varrer as punicoes
                                        c = True

                                        # Vamos obter o ano do guerreiro por aqui mesmo
                                        """Não dava para obter por aqui pois e se a pessoa não tiver punição? Não funcionará."""

                                else:
                                    if c:
                                        # so vai ser executado aqui com o primeiro if
                                        if linha != '}':
                                            # podemos salvar as punições
                                            guerreiro['Punicoes'].append(linha)

                                        else:
                                            # chegamos ao final
                                            break

                pegando_punicoes()

                def ultimo_serv():
                    with open(f'Registros/registro{guerreiro["Turma"]}') as reg:
                        for linha in reg:
                            if '\n' in linha:
                                linha = linha.replace("\n", '')
                            linha = linha.split(',')
                            if linha[0] == guerreiro['Nome Completo']:
                                if linha[1] == '\n' or linha[1] == '':
                                    return 'Sem último serviço'
                                else:
                                    dta = linha[1].split('-')
                                    dta[0], dta[2] = dta[2], dta[0]
                                    return '-'.join(dta)

                guerreiro['Ultimo'] = ultimo_serv()

                # Agora que ja temos todas as informações necessárias, vamos fazer a janela sanhuda

                def criando_ambiente(comp1, alt1):
                    jane_ = Tk()
                    jane_.title("EveryThing")
                    jane_.geometry(f"{comp1}x{alt1}")
                    jane_.configure(bg='#dde')
                    jane_.resizable(False, False)

                    p = Frame(jane_, relief='solid', borderwidth=2, bg='#dde')
                    p.place(x=10, y=10, width=comp1 - 20, height=alt1 - 20)

                    return [jane_, p]

                comp = 400
                alt = 400
                jane, frame = criando_ambiente(comp, alt)

                def exibindo(mestre):
                    opcoes = ['Nome Completo', 'Nome de Guerra', 'ID', 'Seg', 'Serv', 'Turma', 'Ultimo']
                    especificas = ['Nome Completo: ', 'Nome de Guerra: ', 'Número: ', 'Segmento: ',
                                   'Quantidade de Serviços: ', 'Turma: ', 'Último Serviço Registrado: ']

                    b = 10
                    for a in range(0, len(opcoes)):
                        Label(mestre, bg='#dde', text=str(especificas[a]) + str(guerreiro[opcoes[a]])).place(x=10,
                                                                                                             y=10 + a * 30)
                        b += 30

                    Label(mestre, bg='#dde', text='Punições Registradas:').place(x=10, y=b + 30)
                    if len(guerreiro['Punicoes']) != 0:
                        # Agora, devemos mostrar em um combobox, as punições e escalas em que ele aparece
                        punicoes_existentes = ttk.Combobox(mestre, values=guerreiro['Punicoes'])
                        punicoes_existentes['state'] = 'readonly'
                        punicoes_existentes.set(guerreiro['Punicoes'][0])
                        punicoes_existentes.place(x=140, y=b + 30, width=150)
                    else:
                        Label(mestre, bg='#dde', text='Não há punições registradas').place(x=140, y=b + 30)

                    Label(mestre, bg='#dde', text='Escalas Presentes:').place(x=10, y=b + 2 * 30)
                    if len(guerreiro['Escalas']) != 0:
                        escala_existentes = ttk.Combobox(mestre, values=guerreiro['Escalas'])
                        escala_existentes['state'] = 'readonly'
                        escala_existentes.set(guerreiro['Escalas'][0])
                        escala_existentes.place(x=140, y=b + 2 * 30, width=100)
                    else:
                        Label(mestre, bg='#dde', text='Não está presente em escalas').place(x=138, y=b + 2 * 30)

                exibindo(frame)

                def apagando_individual():

                    # Há uma limitação séria, caso o aluno que for apagado seja um específico que possua \n,
                    # o jangau se alastrará e tudo dará errado

                    if messagebox.askyesno(message='Deseja REALMENTE fazer isso?'):
                        messagebox.showinfo(
                            message='É o caso verificar se o apagamento do aluno causou uma modificação destrutiva'
                                    'nos arquivos em que ele estava presente')
                    else:
                        return 0

                    jane.destroy()

                    # Vamos apagar o aluno de todos os pontos

                    def apagando_arquivo_individual(info, arquivo):

                        try:
                            linhas = []
                            presente = False
                            with open(arquivo) as arq:
                                for linha in arq:
                                    if info in linha:
                                        # não vamos salvar essa linha que possui o aluno
                                        presente = True
                                    else:
                                        linhas.append(linha)
                        except FileNotFoundError:
                            return 0

                        if presente:
                            os.remove(arquivo)

                            with open(arquivo, 'x') as arq:
                                for linha_ in linhas:
                                    arq.write(linha_)
                        else:
                            pass

                    apagando_arquivo_individual(guerreiro['ID'], 'SubArquivos/Trocas')
                    apagando_arquivo_individual(guerreiro['ID'], 'SubArquivos/Isentos')
                    apagando_arquivo_individual(guerreiro['ID'], 'SubArquivos/Descansos')
                    apagando_arquivo_individual(guerreiro['Nome de Guerra'], 'SubArquivos/historicobaixados')
                    apagando_arquivo_individual(guerreiro['Nome de Guerra'], 'SubArquivos/baixados')
                    apagando_arquivo_individual(guerreiro['Nome Completo'], f'Registros/registro{guerreiro["Turma"]}')

                    # Para punição, devemos fazer algo diferente.
                    def apagando_punicoes():
                        numero2 = guerreiro['ID']

                        linhas_totais = []
                        linhas_a_serem_apagadas = []
                        achei = False
                        houve_mudanca = False
                        with open(f'Punicoes/punidos{guerreiro["Turma"]}') as base:
                            for linha in base:
                                linhas_totais.append(linha)
                                if numero2 == linha.replace("\n", ''):
                                    # Chegamos no número do sanhudo
                                    achei = True
                                    houve_mudanca = True

                                if achei:
                                    linhas_a_serem_apagadas.append(linha)
                                    if linha == '}\n':
                                        # Chegamos no final
                                        achei = False

                        if houve_mudanca:
                            # Agora, vamos apagar e escrever
                            os.remove(f'Punicoes/punidos{guerreiro["Turma"]}')

                            with open(f'Punicoes/punidos{guerreiro["Turma"]}', 'x') as a_ser_reconstruida:
                                for linha in linhas_totais:
                                    if linha in linhas_a_serem_apagadas:
                                        # não vamos escrever elas
                                        pass
                                    else:
                                        a_ser_reconstruida.write(linha)

                    apagando_punicoes()

                    for esc in os.listdir('../SargenteanteAlpha/Escalas'):
                        apagando_arquivo_individual(guerreiro['ID'], f'Escalas/{esc}')

                    apagando_arquivo_individual(guerreiro['ID'], f'Arquivos/CFG-{guerreiro["Turma"]}.txt')

                Button(jane, text='Apagar Aluno', command=apagando_individual).place(x=25, y=350)

                jane.mainloop()

            def pegando_infor():

                jan = criando_janela("Procurando...", 250, 200)

                Label(jan, text='Insira o tipo de informação:', bg='#dde').place(x=10, y=10)
                tipos_possiveis = ['Número', 'Parte do Nome', 'Parte do Nome de Guerra']
                tipo_ = ttk.Combobox(jan, values=tipos_possiveis)
                tipo_['state'] = 'readonly'
                tipo_.set(tipos_possiveis[1])
                tipo_.place(x=10, y=30, width=150)

                Label(jan, text='Insira a informação:', bg='#dde').place(x=10, y=70)
                inf = Entry(jan)
                inf.place(x=10, y=90)

                def obtendo():
                    tipo = tipo_.get()

                    info = inf.get()

                    # Vamos ter que fazer tudo a montoado mesmo...

                    # Devemos fazer as verificações
                    if tipo == 'Número':
                        if info.isnumeric():
                            tipo = 0
                            jan.destroy()
                        else:
                            messagebox.showerror(message='Não condiz com um número')
                    else:
                        if info.isalpha():
                            if tipo == 'Parte do Nome':
                                tipo = 1
                            else:
                                tipo = 2

                            jan.destroy()
                        else:
                            messagebox.showerror(message='Não condiz com uma palavra')

                    # Agora vamos abrir para procura
                    def comparando(inform, forma, geral):

                        # Há essa diferença, pois em um devemos ter == e no outro apenas in
                        if forma == 0:

                            if inform == geral[0]:
                                return True
                            else:
                                return False
                        else:
                            # Dava ‘bug’ por causa da falta desse upper
                            if inform.upper() in geral[forma]:
                                return True
                            else:
                                return False

                    # Primeiro, uma forma de guardar as informações
                    aluno = {'Nome Completo': [], 'Nome de Guerra': [], 'Seg': [], 'Serv': [], 'ID': [], 'Turma': [],
                             'Punicoes': [],
                             'Escalas': []}
                    pontos = ['ID', 'Nome Completo', 'Nome de Guerra', 'Seg', 'Serv']

                    def extraindo():
                        # Abrindo os arquivos para procura
                        for arquivo in os.listdir('../SargenteanteAlpha/Arquivos'):
                            if '.txt' in arquivo:
                                # Quer dizer que é uma base, logo
                                with open(rf'Arquivos/{arquivo}', 'r') as lista:
                                    for linha_ in lista:
                                        linha = linha_.split("-")

                                        # O sistema deve procurar por toda a lista nelas todas

                                        if linha[0] != '\n':

                                            # Podemos fazer as verificações
                                            if comparando(info, tipo, linha):
                                                linha[4] = linha[4].replace('\n', '')

                                                # Se encontrarmos:
                                                # Uma maneira mais inteligente de fazermos
                                                k = 0
                                                # Não vai dar merda, pois pontos é uma lista limitada
                                                for tema1 in pontos:
                                                    aluno[tema1].append(linha[k])
                                                    k += 1

                                                # Vamos salvar a turma do individuo
                                                for c in range(20, limite_de_busca):
                                                    if f'{c}' in arquivo:
                                                        aluno['Turma'].append(f'{c}')
                                                        break  # ja que não há mais necessidade de continuar

                    extraindo()

                    pontos.append('Turma')
                    """Temos que adicionar só aqui, pois adicionando la estava dando um bug, pois a partir da segunda inter
                    seção, já há opção de turma no pontos, o que da erro, pois a linha não possui essa informação."""

                    # Se tiver mais de um guerreiro, é o caso perguntar
                    if len(aluno['ID']) > 1:
                        def perguntando():
                            ja = Tk()
                            ja.title("Procurando...")
                            ja.geometry("250x200")
                            ja.configure(bg='#dde')
                            ja.resizable(False, False)

                            # Exibindo
                            Label(ja, text='Resultados da busca(ESCOLHA UM):', bg='#dde').place(x=10, y=10)
                            unico = ttk.Combobox(ja, values=aluno['Nome Completo'])
                            unico['state'] = 'readonly'
                            unico.set(aluno['Nome Completo'][0])
                            unico.place(x=10, y=40, width=150)

                            # Pedindo a escolha
                            def escolhendo():
                                # Basicamente vamos limpar toda a lista para sobrar apenas o que foi escolhido
                                nome = unico.get()

                                tam = len(aluno[pontos[1]])
                                for c in range(0, tam):
                                    # Vamos procurá-lo
                                    if aluno[pontos[1]][c] == nome:
                                        # Depois de achá-lo, so precisamos que a lista colapse em um unico termo
                                        for tema7 in pontos:
                                            aluno[tema7] = aluno[tema7][c]
                                        ja.destroy()

                                finalizando(aluno)

                            Button(ja, text='Escolhedor', command=escolhendo).place(x=10, y=80)

                            ja.mainloop()

                        perguntando()

                    else:
                        if len(aluno['ID']) == 0:
                            messagebox.showerror(title='ERROR', message='Aluno Não Encontrado')

                        else:
                            # Como só há um guerreiro, vamos colapsar a lista nele
                            for tema in pontos:
                                aluno[tema] = aluno[tema][0]

                            finalizando(aluno)

                Button(jan, text='Confirmar', command=obtendo).place(x=10, y=140)

                informador(jan, 80, 140, 'A partir de qualquer informação dada, tentará obter as informações do aluno')

                jan.mainloop()

            """Iamos fazer de outra maneira, mas havia um bug de fluxo, logo optamos por fazer desta maneira"""
            pegando_infor()

        # Criando uma função
        barra_menu.add_command(label='Buscador', command=buscador)

        janela.config(menu=barra_menu)

    def isentos():
        # Vamos criar a função que vai abrir uma janela e salvar as pessoas na listas
        def comandos():
            messagebox.showinfo(
                message='Com essa função, você consegue adicionar militares ao grifo ou à turma de comando em vigor')

            janel = criando_janela('', 200, 150)

            def ver_isentos():
                try:
                    janel.destroy()
                except:
                    pass

                # Vamos abrir uma nova janela e bradar as pessoas em um treeview
                jan1 = criando_janela('Isentos Apresentados', 300, 300)

                # Devemos mostrar em um tree view
                def mostrar():

                    def extraindo6():
                        linhas = []
                        with open('../SargenteanteAlpha/SubArquivos/Isentos', 'r') as ise:
                            for linha in ise:
                                if '\n' in linha:
                                    linha = linha.replace("\n", '')
                                linha = linha.split('-')
                                numero3 = pegando_info(linha[0], [0, 2])
                                linhas.append([numero3, linha[1]])

                        return linhas

                    isentos_para_apresentar = extraindo6()

                    tv1 = ttk.Treeview(jan1, columns=['Nome de Guerra', 'Classe'], show='headings')

                    tv1.column('Nome de Guerra', minwidth=0, width=150, anchor='center')
                    tv1.heading('Nome de Guerra', text='Nome de Guerra')

                    tv1.column('Classe', minwidth=0, width=100, anchor='center')
                    tv1.heading('Classe', text='Classe')

                    tv1.place(x=10, y=10, width=280, height=200)

                    for nome, classe in isentos_para_apresentar:
                        tv1.insert('', 'end', values=[nome, classe])

                    return isentos_para_apresentar, tv1

                pessoas, tv = mostrar()

                if len(pessoas) == 0:
                    return messagebox.showwarning(message='Não há pessoas isentas')

                def eliminar():

                    if messagebox.askyesno(message='Tem certeza disso?'):
                        pass
                    else:
                        return 0

                    # Possibilitar apagar alguem
                    pessoa = tv.selection()

                    # noinspection PyTypeChecker
                    pessoa_infos = tv.item(pessoa)['values']

                    # noinspection PyTypeChecker
                    tv.delete(pessoa)

                    # Devemos eliminar do arquivo tbm
                    pessoas.remove(pessoa_infos)

                    os.remove('../SargenteanteAlpha/SubArquivos/Isentos')

                    primeira = True
                    with open('../SargenteanteAlpha/SubArquivos/Isentos', 'x') as ise:
                        for linha in pessoas:
                            linha = '-'.join(linha)
                            if primeira:
                                ise.write(linha)
                                primeira = False
                            else:
                                ise.write(f'\n{linha}')

                    mostrar()

                Button(jan1, text='Eliminar um isento', command=eliminar).place(x=10, y=250)
                informador(jan1, 140, 250, 'Selecione algum isento e retire ele da isenção.')

                jan1.mainloop()

            # Vamos colocar os nomes das funções e respectivas entradas, pedindo para inserir
            def extraindo():

                Label(janel, text='Informe o número: ', bg='#dde').place(x=10, y=20)
                p = Entry(janel)
                p.place(x=10, y=40, width=100)

                Button(janel, text='Ver Lista de Isentos', command=ver_isentos).place(x=10, y=110)

                return p

            entrada = extraindo()

            # QUEM FOR ENTRAR NA TURMA DE COMANDO, DEVE ESTAR SAIR OS ANTERIORES
            def tronando_destronando():

                decisao = messagebox.askyesno(message='Este(a) será do grifo(Sim) ou do comando(Não)')

                num = entrada.get()

                if pegando_info(num, [0, -1]) is None:
                    # não achamos o cara
                    return messagebox.showerror(message='Aluno não encontrado')

                if decisao:
                    decisao = 0
                else:
                    decisao = 1

                def colocando_sanha():
                    possib = ['grifo', 'comando' + str(codigo_mes(DATA_HOJE[-1], False))]

                    # vamos receber as informações que ja existem
                    p = open('../SargenteanteAlpha/SubArquivos/Isentos', 'r')
                    linhas = p.readlines()
                    p.close()

                    # Como ja temos tudo, podemos apenas apagar ele
                    os.remove('../SargenteanteAlpha/SubArquivos/Isentos')

                    def verificacao():
                        for linha3 in linhas:
                            if linha3.split('-')[0] == num:
                                return True

                    deu_ruim = verificacao()
                    if deu_ruim:
                        return messagebox.showwarning(message='Este aluno já está inserido')

                    # Se vamos ter um adicionado ao comando, quer dizer que vai haver a troca, devemos então fazer a limpeza
                    # dos anteriores
                    def limp():
                        c = 0
                        for linha1 in linhas:
                            if 'comando' in linha1:
                                # Devemos analisar
                                linha1 = linha1.split('-')[1].replace('comando', '')
                                if '\n' in linha1:
                                    linha1.replace('\n', '')

                                if linha1 != str(codigo_mes(DATA_HOJE[-1], False)):
                                    linhas.remove(linhas[c])

                            c += 1

                    if decisao == 1:
                        limp()

                    # Depois da limpeza, vamos apenas escrever as informações necessárias
                    if len(linhas) == 0:
                        linhas.append(f'{num}-{possib[decisao]}')
                    else:
                        linhas.append(f'\n{num}-{possib[decisao]}')

                    # Depois salvá-las
                    with open('../SargenteanteAlpha/SubArquivos/Isentos', 'x') as ise:
                        for linha in linhas:
                            ise.write(linha)

                colocando_sanha()

                def verificando_presente_escalas():
                    # Sabendo o número do sanhudo nas escalas

                    for arq in os.listdir('../SargenteanteAlpha/Escalas'):
                        with open(f"Escalas/{arq}", 'r') as esc:
                            for linha in esc:
                                linha = linha.split('-')

                                if linha[0] == num:
                                    messagebox.showwarning(message=f'Este que foi isentado está na escala {arq}')
                                    break

                verificando_presente_escalas()

                janel.destroy()

            Button(janel, text='Tronando-o', command=tronando_destronando).place(x=10, y=70)

            informador(janel, 90, 70, 'Fará com que essa pessoa esteja na turma de comando ou na comunidade do grifo.')

            janel.mainloop()

        barra_menu.add_command(label='Isentos', command=comandos)

        janela.config(menu=barra_menu)

    def apagador():

        # Vamos construir a função principal no subfunções, para podermos usá-la da forma correta.
        # Aqui vamos gerir a busca e la nos preocuparemos em apagar
        def genocidio():
            messagebox.showinfo(
                message='Com essa função, você consegue apagar uma turma inteira e colocar a base na lixeira.')

            # Note que o último arquivo que devemos apagar se trata da base de alunos, óbvio.

            # Devemos perguntar ao usuário de qual turma ele deseja apagar
            def qual_turma():
                jan = criando_janela('Turmas', 220, 140)

                Label(jan, text='Insira a turma que se deseja apagar:', bg='#dde').place(x=10, y=10)
                opcoes_turma = [base for base in os.listdir('../SargenteanteAlpha/Arquivos') if '.txt' in base]
                base_escolhida = ttk.Combobox(jan, values=opcoes_turma)
                base_escolhida['state'] = 'readonly'
                base_escolhida.set(opcoes_turma[0])
                base_escolhida.place(x=10, y=40)

                def obter():
                    base = base_escolhida.get()

                    # ####################################333
                    carater_de_teste = False
                    # ################################3#####

                    if base in opcoes_turma:
                        # ok, existe e vamos usá-la
                        # vamos confirmar
                        if messagebox.askyesno(message='Deseja REALMENTE fazer isso? Cuidado.'):
                            # Vamos fazer o sanha de apagar as coisas

                            # Devemos pegar cada pessoa da base
                            def obtendo_pessoas_gerais():
                                pessoas_turma1 = []

                                with open(f"Arquivos/{base}") as ba:
                                    for linha in ba:
                                        linha = linha.split("-")
                                        if linha[0] != '\n':
                                            if '\n' in linha[-1]:
                                                linha[-1] = linha[-1].replace("\n", '')
                                            pessoas_turma1.append(linha)

                                return pessoas_turma1

                            pessoas_turma = obtendo_pessoas_gerais()

                            numero_turma = 0
                            for i in range(25, limite_de_busca):
                                if str(i) in base:
                                    numero_turma = i
                                    break

                            # Agora, vamos iniciar o processo de apagamento

                            try:
                                # Vamos iniciar o apagamento pelos SubArquivos
                                def apagando_subarquivos():
                                    # Em cada arquivo, há uma forma diferente de organização de informações, por isso precisamos
                                    # pegar informações diferentes da mesma pessoa
                                    subarquivos = ['baixados', 'Descansos', 'historicobaixados', 'Isentos', 'Trocas']
                                    introduzidas = [2, 0, 2, 0, 0]

                                    # Vamos varrer com esta variável
                                    posicao = 0
                                    while True:
                                        if posicao == len(introduzidas):
                                            break

                                        destruidor(
                                            [pessoa_info[introduzidas[posicao]] for pessoa_info in pessoas_turma],
                                            f'SubArquivos/{subarquivos[posicao]}', carater_de_teste)

                                        posicao += 1

                                apagando_subarquivos()

                            except:
                                messagebox.showerror(message='Erro ao tentar apagar alunos nos subarquivos')

                            try:
                                def apagando_registro():
                                    os.remove(f'Registros/registro{numero_turma}')

                                if carater_de_teste:
                                    # nada
                                    pass
                                else:
                                    apagando_registro()

                            except:
                                messagebox.showerror(message='Erro ao apagar registro da turma')

                            try:
                                # Aqui vamos precisar construir de uma forma diferente
                                def apagando_punidos():
                                    os.remove(f'Punicoes/punidos{numero_turma}')

                                if carater_de_teste:
                                    pass
                                else:
                                    apagando_punidos()

                            except:
                                messagebox.showerror(message='Erro ao tentar apagar alunos das punições')

                            try:

                                def apagando_numeros_em_escalas():
                                    numeros = [pessoas_info[0] for pessoas_info in pessoas_turma]

                                    for esc in os.listdir('../SargenteanteAlpha/Escalas'):
                                        destruidor(numeros, f'Escalas/{esc}')

                                if carater_de_teste:
                                    pass
                                else:
                                    apagando_numeros_em_escalas()

                            except:
                                messagebox.showerror(message='Erro ao tentar apagar alunos em escalas')

                            # Por fim,
                            try:
                                if carater_de_teste:
                                    pass
                                else:
                                    os.renames(f'Arquivos/{base}', f'Lixeira/{base}')

                            except:
                                messagebox.showerror(message='Erro ao tentar apagar base de alunos')

                Button(jan, text='Apagá-los', command=obter).place(x=10, y=80)

                jan.mainloop()

            qual_turma()

        barra_menu.add_command(label='Apagador Turma', command=genocidio)

        janela.config(menu=barra_menu)

    def receber_base_dados():
        # Vamos criar uma função capaz de chamar o obtendo alunos mais uma vez

        from TheBigOnes.Sargenteante_Alpha.obtendoalunos import importador

        def iniciar():
            messagebox.showinfo(message='Com essa função, você consegue adicionar um nova turma ao sistema.')
            importador()

        barra_menu.add_command(label='Nova Turma', command=iniciar)

        janela.config(menu=barra_menu)

    mexendo_servicos()
    mexendo_baixados()
    mexendo_punidos()

    buscador_infinito()
    isentos()
    apagador()
    receber_base_dados()

    janela.mainloop()


if __name__ == '__main__':
    sgte()
